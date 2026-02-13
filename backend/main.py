"""
FastAPI Backend for Google Business Scraper
Provides REST API for scraping functionality with authentication
"""

from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends, status, Request, Response, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr
from typing import List, Dict, Optional, Tuple
import pandas as pd
import os
import uuid
import json
import logging
from datetime import datetime, timedelta
from enum import Enum
import asyncio
import sys
import secrets
import hashlib
from urllib.parse import quote_plus
import requests
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
import base64
from io import BytesIO
import config

from payments.coinbase_service import CoinbaseCommerceService
from payments.payment_processor import PaymentProcessor
from payments.pricing_engine import PricingEngine
from payments.webhook_handler import WebhookHandler
from promotions.promo_service import PromoService
from subscriptions.subscription_manager import SubscriptionManager
from audit import record_audit_event

# Fix for Windows asyncio event loop
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from database import init_database, get_db, get_credit_config, CREDITS_STARTING
from auth import (
    verify_password, get_password_hash,
    create_access_token, decode_access_token
)

# ==================== ADMIN SETTINGS UTILITIES ====================

APPROVAL_DENIED = -1
APPROVAL_PENDING = 0
APPROVAL_APPROVED = 1


def _approval_state(value: object) -> int:
    try:
        return int(value)  # type: ignore[arg-type]
    except Exception:
        return APPROVAL_PENDING


def _approval_state_label(value: object) -> str:
    state = _approval_state(value)
    if state == APPROVAL_APPROVED:
        return "approved"
    if state == APPROVAL_DENIED:
        return "denied"
    return "pending"


def get_admin_setting(key: str, default: str = "") -> str:
    """Get admin setting value from database."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM admin_settings WHERE `key` = ?", (key,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else default
    except Exception:
        return default


def set_admin_setting(key: str, value: str, admin_id: int) -> None:
    """Update admin setting in database."""
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    cursor.execute("SELECT 1 FROM admin_settings WHERE `key` = ? LIMIT 1", (key,))
    exists = cursor.fetchone() is not None
    if exists:
        cursor.execute(
            """
            UPDATE admin_settings
            SET value = ?, updated_at = ?, updated_by = ?
            WHERE `key` = ?
            """,
            (value, now, admin_id, key),
        )
    else:
        cursor.execute(
            """
            INSERT INTO admin_settings (`key`, value, updated_at, updated_by)
            VALUES (?, ?, ?, ?)
            """,
            (key, value, now, admin_id),
        )
    conn.commit()
    conn.close()


def get_all_admin_settings() -> dict:
    """Get all admin settings as a dictionary."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT `key`, value FROM admin_settings")
        rows = cursor.fetchall()
        conn.close()
        return {row[0]: row[1] for row in rows}
    except Exception:
        return {}


def _get_starting_credits() -> int:
    raw = get_admin_setting("starting_credits", str(CREDITS_STARTING))
    try:
        credits = int(str(raw).strip())
        return max(0, credits)
    except Exception:
        return int(CREDITS_STARTING)

# ==================== CREDIT SYSTEM UTILITIES ====================

def estimate_credit_cost(num_cities: int, max_results_per_city: int) -> int:
    """Calculate estimated credit cost for a job"""
    credit_config = get_credit_config()
    cost = credit_config["base"] + (num_cities * credit_config["per_city"]) + (num_cities * max_results_per_city * credit_config["per_result"])
    return max(cost, credit_config["min_job"])


def get_user_credit_balance(user_id: int) -> int:
    """Get user's current credit balance"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT credit_balance FROM users WHERE id = ?", (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else 0


def record_credit_transaction(user_id: int, amount: int, balance_after: int,
                              transaction_type: str, reason: str = None,
                              job_id: str = None, created_by: int = None):
    """Record a credit transaction in the ledger"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO credit_ledger (user_id, job_id, amount, balance_after, transaction_type, reason, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (user_id, job_id, amount, balance_after, transaction_type, reason, datetime.now().isoformat(), created_by))
    conn.commit()
    conn.close()


def deduct_credits(user_id: int, amount: int, job_id: str = None, reason: str = None) -> int:
    """Deduct credits from user. Returns new balance."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT credit_balance FROM users WHERE id = ?", (user_id,))
    current = cursor.fetchone()[0] or 0
    new_balance = max(0, current - amount)
    cursor.execute("UPDATE users SET credit_balance = ?, credit_updated_at = ? WHERE id = ?",
                   (new_balance, datetime.now().isoformat(), user_id))
    conn.commit()
    conn.close()
    record_credit_transaction(user_id, -amount, new_balance, "debit", reason or "Job charge", job_id)
    return new_balance


def add_credits(user_id: int, amount: int, reason: str = None, created_by: int = None) -> int:
    """Add credits to user. Returns new balance."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT credit_balance FROM users WHERE id = ?", (user_id,))
    current = cursor.fetchone()[0] or 0
    new_balance = current + amount
    cursor.execute("UPDATE users SET credit_balance = ?, credit_updated_at = ? WHERE id = ?",
                   (new_balance, datetime.now().isoformat(), user_id))
    conn.commit()
    conn.close()
    record_credit_transaction(user_id, amount, new_balance, "credit", reason or "Admin grant", None, created_by)
    return new_balance


def check_rate_limits(user_id: int) -> dict:
    """Check user's rate limits. Returns dict with can_proceed, reason, stats."""
    credit_config = get_credit_config()
    conn = get_db()
    cursor = conn.cursor()

    # Get or create rate limit record
    cursor.execute("SELECT window_start, jobs_in_window FROM rate_limits WHERE user_id = ?", (user_id,))
    rate_record = cursor.fetchone()

    now = datetime.now()
    one_hour_ago = now - timedelta(hours=1)

    if rate_record:
        window_start = datetime.fromisoformat(rate_record[0]) if rate_record[0] else one_hour_ago
        jobs_in_window = rate_record[1] or 0

        # Reset window if more than 1 hour has passed
        if window_start < one_hour_ago:
            cursor.execute("UPDATE rate_limits SET window_start = ?, jobs_in_window = 0 WHERE user_id = ?",
                          (now.isoformat(), user_id))
            conn.commit()
            jobs_in_window = 0
    else:
        # Create new rate limit record
        cursor.execute("INSERT INTO rate_limits (user_id, window_start, jobs_in_window) VALUES (?, ?, ?)",
                      (user_id, now.isoformat(), 0))
        conn.commit()
        jobs_in_window = 0

    # Check concurrent jobs
    cursor.execute("SELECT COUNT(*) FROM jobs WHERE user_id = ? AND status IN ('pending', 'running')", (user_id,))
    concurrent_jobs = cursor.fetchone()[0]

    conn.close()

    if concurrent_jobs >= credit_config["max_concurrent_jobs"]:
        return {
            "can_proceed": False,
            "reason": f"Maximum concurrent jobs ({credit_config['max_concurrent_jobs']}) reached. Wait for current jobs to complete.",
            "concurrent_jobs": concurrent_jobs,
            "jobs_in_hour": jobs_in_window
        }

    if jobs_in_window >= credit_config["max_jobs_per_hour"]:
        return {
            "can_proceed": False,
            "reason": f"Hourly job limit ({credit_config['max_jobs_per_hour']}) reached. Try again later.",
            "concurrent_jobs": concurrent_jobs,
            "jobs_in_hour": jobs_in_window
        }

    return {
        "can_proceed": True,
        "reason": None,
        "concurrent_jobs": concurrent_jobs,
        "jobs_in_hour": jobs_in_window
    }


def increment_rate_limit(user_id: int):
    """Increment user's job count in current rate limit window"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE rate_limits
        SET jobs_in_window = jobs_in_window + 1, last_job_at = ?
        WHERE user_id = ?
    """, (datetime.now().isoformat(), user_id))
    conn.commit()
    conn.close()

# Initialize database
init_database()

app = FastAPI(title="Google Business Scraper API", version="2.0.0", debug=config.DEBUG)

# Audit middleware (attaches request_id, client_ip, user_agent to request.state)
from audit.middleware import AuditMiddleware
app.add_middleware(AuditMiddleware)

@app.get("/")
async def root():
    return {"status": "ok", "service": "Google Business Scraper API"}

@app.get("/favicon.ico")
async def favicon():
    return Response(status_code=204)

# ==================== COUNTRY ENDPOINTS ====================

@app.get("/api/countries")
async def get_countries():
    """Get list of available countries (USA, UK)"""
    return {"countries": list(COUNTRIES_DATA.keys())}

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=config.ALLOWED_ORIGINS,
    allow_origin_regex=config.ALLOWED_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security
security = HTTPBearer()

# Configure logging
logging.basicConfig(
    level=config.LOG_LEVEL,
    format=config.LOG_FORMAT,
    handlers=[
        logging.FileHandler(config.LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load states and cities data for USA
with open(config.USA_DATA_FILE, 'r', encoding='utf-8') as f:
    STATES_CITIES_DATA = json.load(f)

# Load UK regions and cities data
with open(config.UK_DATA_FILE, 'r', encoding='utf-8') as f:
    UK_REGIONS_DATA = json.load(f)

# Load UAE emirates and cities data
with open(config.UAE_DATA_FILE, 'r', encoding='utf-8') as f:
    UAE_CITIES_DATA = json.load(f)

# Load KSA (Saudi Arabia) regions and cities data
with open(config.KSA_DATA_FILE, 'r', encoding='utf-8') as f:
    KSA_CITIES_DATA = json.load(f)

# Load Australia states and cities data
with open(config.AUSTRALIA_DATA_FILE, 'r', encoding='utf-8') as f:
    AUSTRALIA_CITIES_DATA = json.load(f)

# Mapping of country to its region/state data
COUNTRIES_DATA = {
    "USA": STATES_CITIES_DATA,
    "UK": UK_REGIONS_DATA,
    "UAE": UAE_CITIES_DATA,
    "KSA": KSA_CITIES_DATA,
    "Australia": AUSTRALIA_CITIES_DATA
}

# Results directory
RESULTS_DIR = config.RESULTS_DIR
os.makedirs(RESULTS_DIR, exist_ok=True)

# ==================== MODELS ====================

class ScrapingStatus(str, Enum):
    PENDING = "pending"
    RUNNING = "running"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"

class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    phone: Optional[str] = None
    referral_code: Optional[str] = None

class UserLogin(BaseModel):
    username: str
    password: str

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str


class BulkUserIdsRequest(BaseModel):
    user_ids: List[int]


class AdminUserProfileUpdateRequest(BaseModel):
    username: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    is_admin: Optional[bool] = None
    role: Optional[str] = None  # 'user', 'support', 'admin', 'owner'

class ScrapingRequest(BaseModel):
    category: str
    cities_data: List[str]
    max_results_per_city: int = 10

class ScrapingJob(BaseModel):
    job_id: str
    status: ScrapingStatus
    progress: int = 0
    total_cities: int = 0
    current_city: str = ""
    results: List[Dict] = []
    error: Optional[str] = None
    created_at: str
    completed_at: Optional[str] = None
    logs: List[str] = []

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class PurchaseRequest(BaseModel):
    package_id: int
    quantity: int = 1
    promo_code: Optional[str] = None
    provider: str = "stripe"  # 'stripe' or 'coinbase'
    idempotency_key: str


class PromoValidateRequest(BaseModel):
    code: str
    package_id: Optional[int] = None
    quantity: int = 1
    subtotal_cents: Optional[int] = None
    applies_to: str = "packages"  # 'packages' or 'all'


class PriceCalcRequest(BaseModel):
    package_id: int
    quantity: int = 1
    promo_code: Optional[str] = None


class SubscribeRequest(BaseModel):
    plan_id: int
    idempotency_key: str


class AddPaymentMethodRequest(BaseModel):
    stripe_payment_method_id: str
    is_default: bool = False


class PricingTierCreateRequest(BaseModel):
    name: str
    min_monthly_credits: int = 0
    max_monthly_credits: Optional[int] = None
    price_per_credit_cents: float
    discount_percentage: float = 0.0
    requires_approval: bool = False
    is_active: bool = True


class PricingTierUpdateRequest(BaseModel):
    name: Optional[str] = None
    min_monthly_credits: Optional[int] = None
    max_monthly_credits: Optional[int] = None
    price_per_credit_cents: Optional[float] = None
    discount_percentage: Optional[float] = None
    requires_approval: Optional[bool] = None
    is_active: Optional[bool] = None


class PromoCreateRequest(BaseModel):
    code: str
    type: str  # percentage_off|fixed_amount_off|bonus_credits
    discount_percentage: float = 0.0
    discount_amount_cents: int = 0
    bonus_credits: int = 0
    max_uses: Optional[int] = None
    max_uses_per_user: int = 1
    min_purchase_cents: int = 0
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    applies_to: str = "all"
    is_active: bool = True


class PromoUpdateRequest(BaseModel):
    code: Optional[str] = None
    type: Optional[str] = None
    discount_percentage: Optional[float] = None
    discount_amount_cents: Optional[int] = None
    bonus_credits: Optional[int] = None
    max_uses: Optional[int] = None
    max_uses_per_user: Optional[int] = None
    min_purchase_cents: Optional[int] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    applies_to: Optional[str] = None
    is_active: Optional[bool] = None


class CreditPackageCreateRequest(BaseModel):
    name: str
    credits: int
    base_price_cents: int
    display_price_cents: int
    discount_percentage: float = 0.0
    tier_id: Optional[int] = None
    is_active: bool = True
    is_featured: bool = False
    features: Optional[dict] = None


class CreditPackageUpdateRequest(BaseModel):
    name: Optional[str] = None
    credits: Optional[int] = None
    base_price_cents: Optional[int] = None
    display_price_cents: Optional[int] = None
    discount_percentage: Optional[float] = None
    tier_id: Optional[int] = None
    is_active: Optional[bool] = None
    is_featured: Optional[bool] = None
    features: Optional[dict] = None


class SubscriptionPlanCreateRequest(BaseModel):
    name: str
    billing_interval: str  # monthly|yearly
    credits_per_period: int
    base_price_cents: int
    stripe_price_id: Optional[str] = None
    rollover_credits: bool = True
    max_rollover_credits: int = 0
    trial_days: int = 0
    is_active: bool = True
    is_featured: bool = False
    features: Optional[dict] = None


class SubscriptionPlanUpdateRequest(BaseModel):
    name: Optional[str] = None
    billing_interval: Optional[str] = None
    credits_per_period: Optional[int] = None
    base_price_cents: Optional[int] = None
    stripe_price_id: Optional[str] = None
    rollover_credits: Optional[bool] = None
    max_rollover_credits: Optional[int] = None
    trial_days: Optional[int] = None
    is_active: Optional[bool] = None
    is_featured: Optional[bool] = None
    features: Optional[dict] = None

# ==================== AUTHENTICATION ====================

LOGIN_RATE_WINDOW_MINUTES = int(os.getenv("LOGIN_RATE_WINDOW_MINUTES", "10"))
LOGIN_RATE_MAX_ATTEMPTS = int(os.getenv("LOGIN_RATE_MAX_ATTEMPTS", "8"))
LOGIN_RATE_LOCKOUT_MINUTES = int(os.getenv("LOGIN_RATE_LOCKOUT_MINUTES", "10"))


def _get_client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _get_login_rate_limit_record(cursor, scope: str, identifier: str):
    cursor.execute(
        "SELECT id, window_start, attempts_in_window, locked_until FROM login_rate_limits WHERE scope = ? AND identifier = ?",
        (scope, identifier),
    )
    return cursor.fetchone()


def _upsert_login_rate_limit_record(
    cursor,
    record_id,
    scope: str,
    identifier: str,
    window_start: str,
    attempts_in_window: int,
    locked_until: Optional[str],
    last_attempt_at: Optional[str] = None,
    last_success_at: Optional[str] = None,
):
    if record_id is None:
        cursor.execute(
            """
            INSERT INTO login_rate_limits (
                scope, identifier, window_start, attempts_in_window, locked_until, last_attempt_at, last_success_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                scope,
                identifier,
                window_start,
                attempts_in_window,
                locked_until,
                last_attempt_at,
                last_success_at,
            ),
        )
        return

    cursor.execute(
        """
        UPDATE login_rate_limits
        SET window_start = ?,
            attempts_in_window = ?,
            locked_until = ?,
            last_attempt_at = COALESCE(?, last_attempt_at),
            last_success_at = COALESCE(?, last_success_at)
        WHERE id = ?
        """,
        (
            window_start,
            attempts_in_window,
            locked_until,
            last_attempt_at,
            last_success_at,
            record_id,
        ),
    )


def _enforce_login_rate_limit(conn, scope: str, identifier: str):
    """Raise HTTP 429 when the given identifier is currently locked out."""
    cursor = conn.cursor()
    now = datetime.now()
    window_delta = timedelta(minutes=LOGIN_RATE_WINDOW_MINUTES)

    row = _get_login_rate_limit_record(cursor, scope, identifier)
    if not row:
        _upsert_login_rate_limit_record(
            cursor,
            None,
            scope,
            identifier,
            now.isoformat(),
            0,
            None,
        )
        conn.commit()
        return

    record_id, window_start_raw, attempts, locked_until_raw = row
    try:
        window_start = datetime.fromisoformat(window_start_raw) if window_start_raw else now
    except Exception:
        window_start = now

    locked_until = None
    if locked_until_raw:
        try:
            locked_until = datetime.fromisoformat(locked_until_raw)
        except Exception:
            locked_until = None

    if locked_until and locked_until > now:
        retry_after = max(1, int((locked_until - now).total_seconds()))
        raise HTTPException(
            status_code=429,
            detail="Too many login attempts. Please try again later.",
            headers={"Retry-After": str(retry_after)},
        )

    if window_start < now - window_delta:
        _upsert_login_rate_limit_record(
            cursor,
            record_id,
            scope,
            identifier,
            now.isoformat(),
            0,
            None,
        )
        conn.commit()


def _record_login_failure(conn, scope: str, identifier: str) -> bool:
    """
    Record a failed attempt.

    Returns True when the attempt caused a lockout.
    """
    cursor = conn.cursor()
    now = datetime.now()
    window_delta = timedelta(minutes=LOGIN_RATE_WINDOW_MINUTES)
    lock_delta = timedelta(minutes=LOGIN_RATE_LOCKOUT_MINUTES)

    row = _get_login_rate_limit_record(cursor, scope, identifier)
    if row:
        record_id, window_start_raw, attempts, _locked_until_raw = row
        try:
            window_start = datetime.fromisoformat(window_start_raw) if window_start_raw else now
        except Exception:
            window_start = now

        if window_start < now - window_delta:
            attempts = 0
            window_start = now

        attempts = int(attempts or 0) + 1
        locked_until = None
        locked = attempts >= LOGIN_RATE_MAX_ATTEMPTS
        if locked:
            locked_until = (now + lock_delta).isoformat()

        _upsert_login_rate_limit_record(
            cursor,
            record_id,
            scope,
            identifier,
            window_start.isoformat(),
            attempts,
            locked_until,
            last_attempt_at=now.isoformat(),
        )
        conn.commit()
        return locked

    attempts = 1
    locked = attempts >= LOGIN_RATE_MAX_ATTEMPTS
    locked_until = (now + lock_delta).isoformat() if locked else None
    _upsert_login_rate_limit_record(
        cursor,
        None,
        scope,
        identifier,
        now.isoformat(),
        attempts,
        locked_until,
        last_attempt_at=now.isoformat(),
    )
    conn.commit()
    return locked


def _record_login_success(conn, scope: str, identifier: str):
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    row = _get_login_rate_limit_record(cursor, scope, identifier)
    if not row:
        _upsert_login_rate_limit_record(
            cursor,
            None,
            scope,
            identifier,
            now,
            0,
            None,
            last_success_at=now,
        )
        conn.commit()
        return

    record_id = row[0]
    _upsert_login_rate_limit_record(
        cursor,
        record_id,
        scope,
        identifier,
        now,
        0,
        None,
        last_success_at=now,
    )
    conn.commit()


ROLE_HIERARCHY = {"owner": 3, "admin": 2, "support": 1, "user": 0}
VALID_ROLES = set(ROLE_HIERARCHY.keys())

def _user_role_level(user: dict) -> int:
    return ROLE_HIERARCHY.get(user.get("role", "user"), 0)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current authenticated user"""
    token = credentials.credentials
    payload = decode_access_token(token)

    if payload is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials"
        )

    username = payload.get("sub")
    if username is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials"
        )

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, email, phone, is_approved, is_admin, role FROM users WHERE username = ?", (username,))
    user = cursor.fetchone()
    conn.close()

    if user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="User not found"
        )

    approval_state = _approval_state(user[4])
    if approval_state != APPROVAL_APPROVED:
        detail = "Account pending approval"
        if approval_state == APPROVAL_DENIED:
            detail = "Account denied"
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=detail)

    # Derive role: use the role column, fall back to is_admin for backward compat
    role = user[6] if user[6] else ("admin" if bool(user[5]) else "user")

    return {
        "id": user[0],
        "username": user[1],
        "email": user[2],
        "phone": user[3],
        "is_admin": bool(user[5]),
        "role": role,
    }


async def get_admin_user(current_user: dict = Depends(get_current_user)):
    """Require at least 'support' role for read-only admin access."""
    if _user_role_level(current_user) < ROLE_HIERARCHY["support"] and not current_user.get("is_admin"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user


async def get_admin_or_above(current_user: dict = Depends(get_current_user)):
    """Require at least 'admin' role for mutation endpoints."""
    if _user_role_level(current_user) < ROLE_HIERARCHY["admin"] and not current_user.get("is_admin"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin role required"
        )
    return current_user


async def get_owner_user(current_user: dict = Depends(get_current_user)):
    """Require 'owner' role for destructive/sensitive actions."""
    if _user_role_level(current_user) < ROLE_HIERARCHY["owner"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Owner access required"
        )
    return current_user

# ==================== AUTH ENDPOINTS ====================

@app.post("/api/auth/register", status_code=status.HTTP_201_CREATED)
async def register(user_data: UserRegister, request: Request, background_tasks: BackgroundTasks):
    """Register a new user (can be auto-approved if setting enabled)"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if username exists
    cursor.execute("SELECT id FROM users WHERE username = ?", (user_data.username,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email exists
    cursor.execute("SELECT id FROM users WHERE email = ?", (user_data.email,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Check if auto-approve is enabled
    auto_approve = get_admin_setting("auto_approve_users", "false") == "true"
    starting_credits = _get_starting_credits()
    
    # Create user with proper approval status and credits
    password_hash = get_password_hash(user_data.password)
    phone_value = (user_data.phone or "").strip() or None
    created_at = datetime.now().isoformat()
    cursor.execute("""
        INSERT INTO users (username, email, password_hash, phone, is_approved, credit_balance, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_data.username, user_data.email, password_hash, phone_value,
          1 if auto_approve else 0,
          starting_credits if auto_approve else 0,
          created_at))
    
    conn.commit()
    user_id = cursor.lastrowid
    conn.close()

    # Audit: registration
    from audit import record_audit_event
    record_audit_event(
        "auth.register",
        request_id=request.state.request_id,
        actor_user_id=int(user_id),
        actor_username=user_data.username,
        target_type="user",
        target_id=str(user_id),
        after={"username": user_data.username, "email": user_data.email, "auto_approved": auto_approve},
        ip_address=request.state.client_ip,
        user_agent=request.state.user_agent,
    )

    # Optional referral capture
    try:
        from promotions.referral_service import ReferralService
        if user_data.referral_code:
            ReferralService().record_referral_signup(
                referred_user_id=int(user_id),
                referral_code=user_data.referral_code,
            )
    except Exception:
        pass

    if auto_approve:
        # Mark approval metadata + record starting credits in ledger (best effort).
        try:
            conn2 = get_db()
            cursor2 = conn2.cursor()
            cursor2.execute(
                "UPDATE users SET approved_at = ?, approved_by = NULL WHERE id = ?",
                (created_at, user_id),
            )
            conn2.commit()
            conn2.close()
        except Exception:
            pass

        try:
            from promotions.referral_service import ReferralService
            ReferralService().ensure_user_referral_code(int(user_id))
        except Exception:
            pass

        try:
            record_credit_transaction(
                user_id=int(user_id),
                amount=starting_credits,
                balance_after=starting_credits,
                transaction_type="credit",
                reason="Starting credits on auto-approval",
                created_by=None,
            )
        except Exception:
            pass
    else:
        # Notify admins of new pending signup (best effort).
        try:
            if get_admin_setting("admin_notification_on_signup", "true") == "true":
                from notifications.email_service import send_admin_new_signup_email

                recipients: List[str] = []
                try:
                    configured = getattr(config, "ADMIN_NOTIFICATION_EMAILS", None) or []
                    recipients = [e for e in configured if isinstance(e, str) and e.strip()]
                except Exception:
                    recipients = []

                if not recipients:
                    conn3 = get_db()
                    cursor3 = conn3.cursor()
                    cursor3.execute(
                        "SELECT email FROM users WHERE is_admin = 1 AND is_approved = 1 AND email IS NOT NULL AND TRIM(email) != ''"
                    )
                    recipients = [r[0] for r in (cursor3.fetchall() or []) if r and r[0]]
                    conn3.close()

                for to_email in recipients:
                    background_tasks.add_task(
                        send_admin_new_signup_email,
                        to_email=to_email,
                        username=user_data.username,
                        user_email=user_data.email,
                        created_at_iso=created_at,
                    )
        except Exception:
            pass
    
    # Send appropriate email
    try:
        if auto_approve:
            # Send welcome email to auto-approved user
            if get_admin_setting("send_welcome_email", "true") == "true":
                from notifications.email_service import send_welcome_email
                background_tasks.add_task(
                    send_welcome_email,
                    to_email=user_data.email,
                    username=user_data.username,
                    email=user_data.email,
                    starting_credits=starting_credits,
                )
            return {"message": "Registration successful! You can now log in."}
        else:
            return {"message": "Registration successful. Please wait for admin approval."}
    except Exception as e:
        logging.error(f"Error sending welcome email: {e}")
        if auto_approve:
            return {"message": "Registration successful! You can now log in."}
        else:
            return {"message": "Registration successful. Please wait for admin approval."}

@app.post("/api/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin, request: Request):
    """Login and get access token"""
    ip = _get_client_ip(request)
    identifier_username = (credentials.username or "").strip().lower()

    conn = get_db()
    cursor = conn.cursor()

    try:
        _enforce_login_rate_limit(conn, "ip", ip)
        if identifier_username:
            _enforce_login_rate_limit(conn, "username", identifier_username)

        # Allow users to sign in with either their username or their email
        cursor.execute(
            "SELECT id, username, email, phone, password_hash, is_approved, is_admin FROM users WHERE username = ? OR email = ?",
            (credentials.username, credentials.username),
        )
        user = cursor.fetchone()

        if not user or not verify_password(credentials.password, user[4]):
            locked_ip = _record_login_failure(conn, "ip", ip)
            locked_user = False
            if identifier_username:
                locked_user = _record_login_failure(conn, "username", identifier_username)

            # Audit: failed login
            from audit import record_audit_event
            record_audit_event(
                "auth.login",
                request_id=request.state.request_id,
                actor_username=credentials.username,
                target_type="user",
                status="failure",
                failure_reason="Invalid credentials",
                ip_address=request.state.client_ip,
                user_agent=request.state.user_agent,
            )

            if locked_ip or locked_user:
                raise HTTPException(
                    status_code=429,
                    detail="Too many login attempts. Please try again later.",
                )

            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect username or password"
            )

        approval_state = _approval_state(user[5])
        if approval_state != APPROVAL_APPROVED:
            detail = "Account pending approval"
            if approval_state == APPROVAL_DENIED:
                detail = "Account denied"
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=detail)

        _record_login_success(conn, "ip", ip)
        if identifier_username:
            _record_login_success(conn, "username", identifier_username)

        cursor.execute("UPDATE users SET last_login = ? WHERE id = ?", (datetime.now().isoformat(), user[0]))
        conn.commit()
    finally:
        conn.close()
    
    # Audit: successful login
    from audit import record_audit_event
    record_audit_event(
        "auth.login",
        request_id=request.state.request_id,
        actor_user_id=user[0],
        actor_username=user[1],
        target_type="user",
        target_id=str(user[0]),
        ip_address=request.state.client_ip,
        user_agent=request.state.user_agent,
    )

    # Create token
    access_token = create_access_token(data={"sub": user[1]})

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user[0],
            "username": user[1],
            "email": user[2],
            "phone": user[3],
            "is_admin": bool(user[6])
        }
    }

@app.get("/api/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    return current_user

@app.post("/api/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Request password reset token"""
    conn = get_db()
    cursor = conn.cursor()

    # Check if email exists
    cursor.execute("SELECT id, username FROM users WHERE email = ?", (request.email,))
    user = cursor.fetchone()

    # Always return success message (security best practice - don't reveal if email exists)
    if not user:
        conn.close()
        return {"message": "If that email is registered, a password reset link has been sent."}

    user_id = user[0]

    # Generate secure random token
    reset_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(reset_token.encode()).hexdigest()

    # Token expires in 24 hours
    expires_at = (datetime.now() + timedelta(hours=24)).isoformat()

    # Store token in database
    cursor.execute("""
        INSERT INTO password_reset_tokens (user_id, token_hash, expires_at, created_at)
        VALUES (?, ?, ?, ?)
    """, (user_id, token_hash, expires_at, datetime.now().isoformat()))
    conn.commit()
    conn.close()

    # TODO: Send email with reset link
    # For MVP, log the token to console
    reset_url = f"http://localhost:8000/reset-password/{reset_token}"
    print(f"\n{'='*60}")
    print(f"PASSWORD RESET TOKEN FOR: {request.email}")
    print(f"Reset URL: {reset_url}")
    print(f"Token: {reset_token}")
    print(f"{'='*60}\n")

    return {"message": "If that email is registered, a password reset link has been sent."}

@app.get("/api/auth/verify-reset-token/{token}")
async def verify_reset_token(token: str):
    """Verify if reset token is valid"""
    token_hash = hashlib.sha256(token.encode()).hexdigest()

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT rt.id, rt.user_id, rt.expires_at, rt.used, u.email
        FROM password_reset_tokens rt
        JOIN users u ON rt.user_id = u.id
        WHERE rt.token_hash = ?
    """, (token_hash,))

    token_data = cursor.fetchone()
    conn.close()

    if not token_data:
        raise HTTPException(status_code=400, detail="Invalid reset token")

    token_id, user_id, expires_at, used, email = token_data

    if used:
        raise HTTPException(status_code=400, detail="Reset token has already been used")

    if datetime.fromisoformat(expires_at) < datetime.now():
        raise HTTPException(status_code=400, detail="Reset token has expired")

    return {"valid": True, "email": email}

@app.post("/api/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset password using token"""
    token_hash = hashlib.sha256(request.token.encode()).hexdigest()

    conn = get_db()
    cursor = conn.cursor()

    # Verify token
    cursor.execute("""
        SELECT id, user_id, expires_at, used
        FROM password_reset_tokens
        WHERE token_hash = ?
    """, (token_hash,))

    token_data = cursor.fetchone()

    if not token_data:
        conn.close()
        raise HTTPException(status_code=400, detail="Invalid reset token")

    token_id, user_id, expires_at, used = token_data

    if used:
        conn.close()
        raise HTTPException(status_code=400, detail="Reset token has already been used")

    if datetime.fromisoformat(expires_at) < datetime.now():
        conn.close()
        raise HTTPException(status_code=400, detail="Reset token has expired")

    # Update password
    new_password_hash = get_password_hash(request.new_password)
    cursor.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_password_hash, user_id))

    # Mark token as used
    cursor.execute("UPDATE password_reset_tokens SET used = 1, used_at = ? WHERE id = ?",
                   (datetime.now().isoformat(), token_id))

    conn.commit()
    conn.close()

    return {"message": "Password reset successful"}


# Debug endpoint: list users (local dev only)
@app.get("/internal/debug_users")
async def _debug_list_users():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, username, email, phone, is_approved, is_admin, created_at, last_login, approved_at, denied_at, denied_reason FROM users ORDER BY id DESC"
    )
    users = cursor.fetchall()
    conn.close()
    return {
        "users": [
            {
                "id": u[0],
                "username": u[1],
                "email": u[2],
                "phone": u[3],
                "approval_status": _approval_state(u[4]),
                "approval_state": _approval_state_label(u[4]),
                "is_approved": _approval_state(u[4]) == APPROVAL_APPROVED,
                "is_admin": bool(u[5]),
                "created_at": u[6],
                "last_login": u[7],
                "approved_at": u[8],
                "denied_at": u[9],
                "denied_reason": u[10],
            }
            for u in users
        ]
    }

# ==================== ADMIN ENDPOINTS ====================

@app.get("/api/admin/users")
async def get_all_users(admin: dict = Depends(get_admin_user)):
    """Get all users (admin only)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, username, email, phone, is_approved, is_admin, credit_balance, created_at, last_login,
               approved_at, approved_by, denied_at, denied_by, denied_reason, role
        FROM users ORDER BY created_at DESC
    """)
    users = cursor.fetchall()
    conn.close()

    return {
        "users": [
            {
                "id": u[0],
                "username": u[1],
                "email": u[2],
                "phone": u[3],
                "approval_status": _approval_state(u[4]),
                "approval_state": _approval_state_label(u[4]),
                "is_approved": _approval_state(u[4]) == APPROVAL_APPROVED,
                "is_denied": _approval_state(u[4]) == APPROVAL_DENIED,
                "is_admin": bool(u[5]),
                "credit_balance": u[6] or 0,
                "created_at": u[7],
                "last_login": u[8],
                "approved_at": u[9],
                "approved_by": u[10],
                "denied_at": u[11],
                "denied_by": u[12],
                "denied_reason": u[13],
                "role": u[14] or ("admin" if bool(u[5]) else "user"),
            }
            for u in users
        ]
    }


@app.put("/api/admin/users/{user_id}")
async def update_user_profile(
    user_id: int,
    profile_data: AdminUserProfileUpdateRequest,
    request: Request,
    admin: dict = Depends(get_admin_user),
):
    """Update user profile fields (admin only)."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT id, username, email, phone, is_admin FROM users WHERE id = ?",
        (user_id,),
    )
    existing = cursor.fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")

    before_state = {
        "username": existing[1],
        "email": existing[2],
        "phone": existing[3],
        "is_admin": bool(existing[4]),
    }

    update_fields: List[str] = []
    params: List[object] = []

    if profile_data.username is not None:
        username = profile_data.username.strip()
        if len(username) < 3:
            conn.close()
            raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
        cursor.execute(
            "SELECT id FROM users WHERE username = ? AND id != ?",
            (username, user_id),
        )
        if cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=400, detail="Username already exists")
        update_fields.append("username = ?")
        params.append(username)

    if profile_data.email is not None:
        email = str(profile_data.email).strip().lower()
        cursor.execute(
            "SELECT id FROM users WHERE email = ? AND id != ?",
            (email, user_id),
        )
        if cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=400, detail="Email already exists")
        update_fields.append("email = ?")
        params.append(email)

    if profile_data.phone is not None:
        phone = profile_data.phone.strip()
        update_fields.append("phone = ?")
        params.append(phone or None)

    if profile_data.role is not None:
        role = profile_data.role.strip().lower()
        if role not in VALID_ROLES:
            conn.close()
            raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(VALID_ROLES)}")
        # Prevent demoting the last owner/admin
        cursor.execute(
            "SELECT role FROM users WHERE id = ?", (user_id,),
        )
        current_role_row = cursor.fetchone()
        current_role = current_role_row[0] if current_role_row and current_role_row[0] else "user"
        if ROLE_HIERARCHY.get(current_role, 0) >= ROLE_HIERARCHY["admin"] and ROLE_HIERARCHY.get(role, 0) < ROLE_HIERARCHY["admin"]:
            cursor.execute("SELECT COUNT(*) FROM users WHERE role IN ('admin', 'owner') OR is_admin = 1")
            admin_count = int((cursor.fetchone() or [0])[0] or 0)
            if admin_count <= 1:
                conn.close()
                raise HTTPException(status_code=400, detail="Cannot remove the last admin/owner")
        update_fields.append("role = ?")
        params.append(role)
        # Keep is_admin in sync
        update_fields.append("is_admin = ?")
        params.append(1 if role in ("admin", "owner", "support") else 0)

    elif profile_data.is_admin is not None:
        target_is_admin = bool(existing[4])
        if target_is_admin and not profile_data.is_admin:
            cursor.execute("SELECT COUNT(*) FROM users WHERE is_admin = 1")
            admin_count = int((cursor.fetchone() or [0])[0] or 0)
            if admin_count <= 1:
                conn.close()
                raise HTTPException(status_code=400, detail="Cannot remove the last admin")
        update_fields.append("is_admin = ?")
        params.append(1 if profile_data.is_admin else 0)
        # Keep role in sync
        update_fields.append("role = ?")
        params.append("admin" if profile_data.is_admin else "user")

    if not update_fields:
        conn.close()
        raise HTTPException(status_code=400, detail="No profile fields to update")

    params.append(user_id)
    cursor.execute(
        f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?",
        tuple(params),
    )
    conn.commit()
    conn.close()

    after_state = {k: v for k, v in profile_data.dict(exclude_unset=True).items()}

    record_audit_event(
        "admin.user.update",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        before=before_state,
        after=after_state,
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
    )

    return {"ok": True}

@app.post("/api/admin/users/{user_id}/approve")
async def approve_user(
    user_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(get_admin_user),
):
    """Approve a user and grant starting credits (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    # Check if user exists and is not already approved
    cursor.execute("SELECT is_approved, credit_balance, email, username FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")

    approval_state = _approval_state(user[0])
    if approval_state == APPROVAL_APPROVED:
        conn.close()
        return {"message": "User already approved"}
    if approval_state == APPROVAL_DENIED:
        conn.close()
        raise HTTPException(status_code=400, detail="User is denied. Restore the user before approving.")

    user_email = user[2]
    username = user[3]
    current_balance = int(user[1] or 0)
    starting_credits = _get_starting_credits()
    new_balance = current_balance + starting_credits
    now = datetime.now().isoformat()

    # Approve user and grant starting credits
    cursor.execute("""
        UPDATE users
        SET is_approved = 1,
            credit_balance = ?,
            credit_updated_at = ?,
            approved_at = ?,
            approved_by = ?,
            denied_at = NULL,
            denied_by = NULL,
            denied_reason = NULL
        WHERE id = ?
    """, (new_balance, now, now, admin["id"], user_id))
    conn.commit()
    conn.close()

    # Ensure user has a referral code for sharing
    try:
        from promotions.referral_service import ReferralService
        ReferralService().ensure_user_referral_code(user_id)
    except Exception:
        pass

    # Record the credit grant in ledger
    record_credit_transaction(
        user_id=user_id,
        amount=starting_credits,
        balance_after=new_balance,
        transaction_type="credit",
        reason="Starting credits on account approval",
        created_by=admin["id"]
    )

    # Send approval email if enabled
    try:
        if get_admin_setting("send_approval_email", "true") == "true" and user_email:
            from notifications.email_service import send_approval_email
            background_tasks.add_task(
                send_approval_email,
                to_email=user_email,
                username=username,
                starting_credits=starting_credits,
            )
    except Exception as e:
        logging.error(f"Error sending approval email: {e}")

    record_audit_event(
        "admin.user.approve",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        before={"is_approved": approval_state},
        after={"is_approved": 1},
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
    )

    return {"message": f"User approved successfully with {starting_credits} starting credits"}


@app.post("/api/admin/users/bulk-approve")
async def bulk_approve_users(
    payload: BulkUserIdsRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(get_admin_user),
):
    """Bulk-approve users (admin only)."""
    user_ids = sorted({int(u) for u in (payload.user_ids or []) if int(u) > 0})
    if not user_ids:
        return {"approved": 0, "skipped": 0, "not_found": 0, "user_ids": []}

    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    starting_credits = _get_starting_credits()

    approved_ids: List[int] = []
    skipped_ids: List[int] = []
    not_found_ids: List[int] = []
    approved_infos: List[Tuple[int, Optional[str], Optional[str], int]] = []

    try:
        for user_id in user_ids:
            cursor.execute("SELECT is_approved, credit_balance, email, username FROM users WHERE id = ?", (user_id,))
            row = cursor.fetchone()
            if not row:
                not_found_ids.append(user_id)
                continue

            state = _approval_state(row[0])
            if state != APPROVAL_PENDING:
                skipped_ids.append(user_id)
                continue

            current_balance = int(row[1] or 0)
            new_balance = current_balance + starting_credits

            cursor.execute(
                """
                UPDATE users
                SET is_approved = 1,
                    credit_balance = ?,
                    credit_updated_at = ?,
                    approved_at = ?,
                    approved_by = ?,
                    denied_at = NULL,
                    denied_by = NULL,
                    denied_reason = NULL
                WHERE id = ?
                """,
                (new_balance, now, now, admin["id"], user_id),
            )
            approved_ids.append(user_id)
            approved_infos.append((user_id, row[2], row[3], new_balance))

        conn.commit()
    finally:
        conn.close()

    # Best-effort post-approval steps (ledger + referral code)
    for user_id, user_email, username, new_balance in approved_infos:
        try:
            from promotions.referral_service import ReferralService
            ReferralService().ensure_user_referral_code(user_id)
        except Exception:
            pass

        try:
            record_credit_transaction(
                user_id=user_id,
                amount=starting_credits,
                balance_after=new_balance,
                transaction_type="credit",
                reason="Starting credits on account approval (bulk)",
                created_by=admin["id"],
            )
        except Exception:
            pass

        try:
            if get_admin_setting("send_approval_email", "true") == "true" and user_email and username:
                from notifications.email_service import send_approval_email
                background_tasks.add_task(
                    send_approval_email,
                    to_email=user_email,
                    username=username,
                    starting_credits=starting_credits,
                )
        except Exception:
            pass

    record_audit_event(
        "admin.user.bulk_approve",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
        metadata={"user_ids": approved_ids},
    )

    return {
        "approved": len(approved_ids),
        "skipped": len(skipped_ids),
        "not_found": len(not_found_ids),
        "user_ids": approved_ids,
    }

@app.post("/api/admin/users/{user_id}/deny")
async def deny_user(
    user_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    admin_note: Optional[str] = None,
    admin: dict = Depends(get_admin_user),
):
    """Deny/reject a user registration (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    # Check if user exists and is not approved
    cursor.execute("SELECT is_approved, email, username FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")

    approval_state = _approval_state(user[0])
    if approval_state == APPROVAL_APPROVED:
        conn.close()
        raise HTTPException(status_code=400, detail="User is already approved")
    if approval_state == APPROVAL_DENIED:
        conn.close()
        return {"message": "User is already denied"}

    user_email = user[1]
    username = user[2]

    # Delete or mark as denied (we'll mark them with is_denied flag, or delete)
    # For now, let's soft delete by setting is_approved to -1 (denied state)
    cursor.execute("""
        UPDATE users
        SET is_approved = -1,
            denied_at = ?,
            denied_by = ?,
            denied_reason = ?,
            approved_at = NULL,
            approved_by = NULL
        WHERE id = ?
    """, (datetime.now().isoformat(), admin["id"], (admin_note or None), user_id))
    conn.commit()
    conn.close()

    # Send rejection email if enabled
    try:
        if get_admin_setting("send_rejection_email", "true") == "true" and user_email:
            from notifications.email_service import send_rejection_email
            background_tasks.add_task(
                send_rejection_email,
                to_email=user_email,
                username=username,
                admin_note=admin_note,
            )
    except Exception as e:
        logging.error(f"Error sending rejection email: {e}")

    record_audit_event(
        "admin.user.deny",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        before={"is_approved": approval_state},
        after={"is_approved": -1},
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
    )

    return {"message": "User registration denied"}


@app.post("/api/admin/users/{user_id}/restore")
async def restore_user(user_id: int, request: Request, admin: dict = Depends(get_admin_user)):
    """Restore a denied user back to pending status (admin only)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT is_approved FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")

    state = _approval_state(row[0])
    if state != APPROVAL_DENIED:
        conn.close()
        raise HTTPException(status_code=400, detail="User is not denied")

    cursor.execute(
        """
        UPDATE users
        SET is_approved = 0,
            denied_at = NULL,
            denied_by = NULL,
            denied_reason = NULL
        WHERE id = ?
        """,
        (user_id,),
    )
    conn.commit()
    conn.close()
    record_audit_event(
        "admin.user.restore",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        before={"is_approved": APPROVAL_DENIED},
        after={"is_approved": 0},
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
    )

    return {"message": "User restored to pending approval"}

@app.delete("/api/admin/users/{user_id}")
async def delete_user(user_id: int, request: Request, admin: dict = Depends(get_admin_user)):
    """Delete a user (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, username, email, is_admin FROM users WHERE id = ?", (user_id,))
    existing = cursor.fetchone()
    before_state = None
    if existing:
        before_state = {
            "id": existing[0],
            "username": existing[1],
            "email": existing[2],
            "is_admin": bool(existing[3]),
        }

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()

    record_audit_event(
        "admin.user.delete",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        before=before_state,
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
    )

    return {"message": "User deleted successfully"}

@app.get("/api/admin/stats")
async def get_admin_stats(admin: dict = Depends(get_admin_user)):
    """Get admin dashboard statistics"""
    conn = get_db()
    cursor = conn.cursor()

    # Total users
    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]

    # Approved users
    cursor.execute("SELECT COUNT(*) FROM users WHERE is_approved = 1")
    approved_users = cursor.fetchone()[0]

    # Pending users
    cursor.execute("SELECT COUNT(*) FROM users WHERE is_approved = 0")
    pending_users = cursor.fetchone()[0]

    # Total jobs
    cursor.execute("SELECT COUNT(*) FROM jobs")
    total_jobs = cursor.fetchone()[0]

    # Completed jobs
    cursor.execute("SELECT COUNT(*) FROM jobs WHERE status = 'completed'")
    completed_jobs = cursor.fetchone()[0]

    # Total results
    cursor.execute("SELECT COUNT(*) FROM results")
    total_results = cursor.fetchone()[0]

    # Total credits in circulation
    cursor.execute("SELECT SUM(credit_balance) FROM users")
    total_credits = cursor.fetchone()[0] or 0

    # Pending credit requests
    cursor.execute("SELECT COUNT(*) FROM credit_requests WHERE status = 'pending'")
    pending_credit_requests = cursor.fetchone()[0]

    # Recent jobs
    cursor.execute("""
        SELECT j.job_id, j.category, j.status, j.created_at, u.username
        FROM jobs j
        JOIN users u ON j.user_id = u.id
        ORDER BY j.created_at DESC
        LIMIT 10
    """)
    recent_jobs = [
        {
            "job_id": j[0],
            "category": j[1],
            "status": j[2],
            "created_at": j[3],
            "username": j[4]
        }
        for j in cursor.fetchall()
    ]

    conn.close()

    return {
        "total_users": total_users,
        "approved_users": approved_users,
        "pending_users": pending_users,
        "total_jobs": total_jobs,
        "completed_jobs": completed_jobs,
        "total_results": total_results,
        "total_credits": total_credits,
        "pending_credit_requests": pending_credit_requests,
        "recent_jobs": recent_jobs
    }


# ==================== ADMIN SETTINGS ENDPOINTS ====================

@app.get("/api/admin/settings")
async def get_all_settings(admin: dict = Depends(get_admin_user)):
    """Get all admin settings (admin only)"""
    return get_all_admin_settings()


@app.get("/api/admin/settings/{key}")
async def get_setting(key: str, admin: dict = Depends(get_admin_user)):
    """Get a specific admin setting (admin only)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM admin_settings WHERE `key` = ?", (key,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")
    return {"key": key, "value": row[0]}


class AdminSettingUpdate(BaseModel):
    value: str


@app.put("/api/admin/settings/{key}")
async def update_setting(
    key: str,
    req: Request,
    payload: Optional[AdminSettingUpdate] = Body(default=None),
    value: Optional[str] = None,
    admin: dict = Depends(get_admin_user),
):
    """Update an admin setting (admin only)"""
    new_value = payload.value if payload is not None else value
    if new_value is None:
        raise HTTPException(status_code=400, detail="Missing setting value")
    old_value = get_admin_setting(key)
    set_admin_setting(key, str(new_value), admin["id"])

    record_audit_event(
        "admin.setting.update",
        request_id=getattr(req.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="setting",
        target_id=key,
        before={"value": old_value},
        after={"value": str(new_value)},
        ip_address=getattr(req.state, "client_ip", None),
        user_agent=getattr(req.state, "user_agent", None),
    )

    return {
        "message": f"Setting '{key}' updated successfully",
        "key": key,
        "value": str(new_value),
    }


# ==================== CREDIT ENDPOINTS ====================

class CreditEstimateRequest(BaseModel):
    num_cities: int
    max_results_per_city: int = 10


class CreditRequestModel(BaseModel):
    amount_requested: int
    reason: Optional[str] = None


class AdminCreditGrant(BaseModel):
    amount: int
    reason: Optional[str] = None


class BulkCreditRequestIdsRequest(BaseModel):
    request_ids: List[int]
    admin_note: Optional[str] = None


@app.get("/api/credits/config")
async def get_credits_config(current_user: dict = Depends(get_current_user)):
    """Get credit system configuration"""
    return get_credit_config()


@app.get("/api/credits/balance")
async def get_credit_balance(current_user: dict = Depends(get_current_user)):
    """Get current user's credit balance and rate limit status"""
    balance = get_user_credit_balance(current_user["id"])
    rate_status = check_rate_limits(current_user["id"])
    config = get_credit_config()

    return {
        "balance": balance,
        "rate_limits": {
            "max_jobs_per_hour": config["max_jobs_per_hour"],
            "max_concurrent_jobs": config["max_concurrent_jobs"],
            "jobs_in_hour": rate_status["jobs_in_hour"],
            "concurrent_jobs": rate_status["concurrent_jobs"],
            "can_create_job": rate_status["can_proceed"]
        }
    }


@app.post("/api/credits/estimate")
async def estimate_job_cost(request: CreditEstimateRequest, current_user: dict = Depends(get_current_user)):
    """Estimate credit cost for a job before submission"""
    cost = estimate_credit_cost(request.num_cities, request.max_results_per_city)
    balance = get_user_credit_balance(current_user["id"])

    return {
        "estimated_cost": cost,
        "current_balance": balance,
        "sufficient_credits": balance >= cost,
        "balance_after": balance - cost if balance >= cost else 0
    }


@app.get("/api/credits/history")
async def get_credit_history(current_user: dict = Depends(get_current_user)):
    """Get user's credit transaction history"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, job_id, amount, balance_after, transaction_type, reason, created_at
        FROM credit_ledger
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 50
    """, (current_user["id"],))

    transactions = [
        {
            "id": row[0],
            "job_id": row[1],
            "amount": row[2],
            "balance_after": row[3],
            "transaction_type": row[4],
            "reason": row[5],
            "created_at": row[6]
        }
        for row in cursor.fetchall()
    ]
    conn.close()

    return {"transactions": transactions}


@app.post("/api/credits/request")
async def request_credits(request: CreditRequestModel, current_user: dict = Depends(get_current_user)):
    """Submit a request for more credits"""
    conn = get_db()
    cursor = conn.cursor()

    # Check for pending requests
    cursor.execute("SELECT COUNT(*) FROM credit_requests WHERE user_id = ? AND status = 'pending'", (current_user["id"],))
    pending = cursor.fetchone()[0]
    if pending > 0:
        conn.close()
        raise HTTPException(status_code=400, detail="You already have a pending credit request")

    cursor.execute("""
        INSERT INTO credit_requests (user_id, amount_requested, reason, created_at)
        VALUES (?, ?, ?, ?)
    """, (current_user["id"], request.amount_requested, request.reason, datetime.now().isoformat()))
    conn.commit()
    conn.close()

    return {"message": "Credit request submitted successfully"}


@app.get("/api/credits/requests")
async def get_my_credit_requests(current_user: dict = Depends(get_current_user)):
    """Get user's credit request history"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, amount_requested, reason, status, admin_note, created_at, reviewed_at
        FROM credit_requests
        WHERE user_id = ?
        ORDER BY created_at DESC
    """, (current_user["id"],))

    requests = [
        {
            "id": row[0],
            "amount_requested": row[1],
            "reason": row[2],
            "status": row[3],
            "admin_note": row[4],
            "created_at": row[5],
            "reviewed_at": row[6]
        }
        for row in cursor.fetchall()
    ]
    conn.close()

    return {"requests": requests}


# ==================== ADMIN CREDIT MANAGEMENT ====================

@app.get("/api/admin/credits/requests")
async def get_all_credit_requests(admin: dict = Depends(get_admin_user)):
    """Get all pending credit requests (admin only)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT cr.id, cr.user_id, u.username, u.email, cr.amount_requested, cr.reason, cr.status, cr.created_at
        FROM credit_requests cr
        JOIN users u ON cr.user_id = u.id
        WHERE cr.status = 'pending'
        ORDER BY cr.created_at ASC
    """)

    requests = [
        {
            "id": row[0],
            "user_id": row[1],
            "username": row[2],
            "email": row[3],
            "amount_requested": row[4],
            "reason": row[5],
            "status": row[6],
            "created_at": row[7]
        }
        for row in cursor.fetchall()
    ]
    conn.close()

    return {"requests": requests}


@app.post("/api/admin/credits/requests/{request_id}/approve")
async def approve_credit_request(request_id: int, req: Request, admin: dict = Depends(get_admin_user)):
    """Approve a credit request (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    # Get request details
    cursor.execute("SELECT user_id, amount_requested FROM credit_requests WHERE id = ? AND status = 'pending'", (request_id,))
    cr = cursor.fetchone()
    if not cr:
        conn.close()
        raise HTTPException(status_code=404, detail="Credit request not found or already processed")

    user_id, amount = cr

    # Update request status
    cursor.execute("""
        UPDATE credit_requests SET status = 'approved', reviewed_by = ?, reviewed_at = ?
        WHERE id = ?
    """, (admin["id"], datetime.now().isoformat(), request_id))
    conn.commit()
    conn.close()

    # Add credits to user
    new_balance = add_credits(user_id, amount, "Credit request approved", admin["id"])

    record_audit_event(
        "admin.credits.request_approve",
        request_id=getattr(req.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="credit_request",
        target_id=str(request_id),
        ip_address=getattr(req.state, "client_ip", None),
        user_agent=getattr(req.state, "user_agent", None),
        metadata={"user_id": user_id, "amount": amount},
    )

    return {"message": "Credit request approved", "new_balance": new_balance}


@app.post("/api/admin/credits/requests/bulk-approve")
async def bulk_approve_credit_requests(payload: BulkCreditRequestIdsRequest, request: Request, admin: dict = Depends(get_admin_user)):
    """Bulk-approve credit requests (admin only)."""
    request_ids = sorted({int(r) for r in (payload.request_ids or []) if int(r) > 0})
    if not request_ids:
        return {"approved": 0, "skipped": 0, "not_found": 0, "request_ids": []}

    approved_ids: List[int] = []
    skipped_ids: List[int] = []
    not_found_ids: List[int] = []

    for request_id in request_ids:
        conn = get_db()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "SELECT user_id, amount_requested FROM credit_requests WHERE id = ? AND status = 'pending'",
                (request_id,),
            )
            req = cursor.fetchone()
            if not req:
                cursor.execute("SELECT 1 FROM credit_requests WHERE id = ?", (request_id,))
                exists = cursor.fetchone()
                if exists:
                    skipped_ids.append(request_id)
                else:
                    not_found_ids.append(request_id)
                continue

            user_id, amount = req
            cursor.execute(
                """
                UPDATE credit_requests
                SET status = 'approved', reviewed_by = ?, reviewed_at = ?
                WHERE id = ?
                """,
                (admin["id"], datetime.now().isoformat(), request_id),
            )
            conn.commit()
        finally:
            conn.close()

        try:
            add_credits(int(user_id), int(amount), "Credit request approved (bulk)", admin["id"])
        except Exception:
            # Request is already approved; keep going (matches single-endpoint semantics).
            pass

        approved_ids.append(request_id)

    if approved_ids:
        record_audit_event(
            "admin.credits.bulk_approve", request_id=request.state.request_id,
            actor_user_id=admin["id"], actor_username=admin["username"],
            target_type="credit_request",
            metadata={"approved_ids": approved_ids, "count": len(approved_ids)},
            ip_address=request.state.client_ip, user_agent=request.state.user_agent,
        )

    return {
        "approved": len(approved_ids),
        "skipped": len(skipped_ids),
        "not_found": len(not_found_ids),
        "request_ids": approved_ids,
    }


@app.post("/api/admin/credits/requests/{request_id}/deny")
async def deny_credit_request(request_id: int, req: Request, admin_note: Optional[str] = None, admin: dict = Depends(get_admin_user)):
    """Deny a credit request (admin only)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE credit_requests SET status = 'denied', admin_note = ?, reviewed_by = ?, reviewed_at = ?
        WHERE id = ? AND status = 'pending'
    """, (admin_note, admin["id"], datetime.now().isoformat(), request_id))

    if cursor.rowcount == 0:
        conn.close()
        raise HTTPException(status_code=404, detail="Credit request not found or already processed")

    conn.commit()
    conn.close()

    record_audit_event(
        "admin.credits.request_deny",
        request_id=getattr(req.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="credit_request",
        target_id=str(request_id),
        ip_address=getattr(req.state, "client_ip", None),
        user_agent=getattr(req.state, "user_agent", None),
    )

    return {"message": "Credit request denied"}


@app.post("/api/admin/credits/requests/bulk-deny")
async def bulk_deny_credit_requests(payload: BulkCreditRequestIdsRequest, request: Request, admin: dict = Depends(get_admin_user)):
    """Bulk-deny credit requests (admin only)."""
    request_ids = sorted({int(r) for r in (payload.request_ids or []) if int(r) > 0})
    if not request_ids:
        return {"denied": 0, "skipped": 0, "not_found": 0, "request_ids": []}

    conn = get_db()
    cursor = conn.cursor()
    denied_ids: List[int] = []
    skipped_ids: List[int] = []
    not_found_ids: List[int] = []
    now = datetime.now().isoformat()

    try:
        for request_id in request_ids:
            cursor.execute(
                "UPDATE credit_requests SET status = 'denied', admin_note = ?, reviewed_by = ?, reviewed_at = ? WHERE id = ? AND status = 'pending'",
                (payload.admin_note, admin["id"], now, request_id),
            )
            if cursor.rowcount > 0:
                denied_ids.append(request_id)
                continue

            cursor.execute("SELECT 1 FROM credit_requests WHERE id = ?", (request_id,))
            exists = cursor.fetchone()
            if exists:
                skipped_ids.append(request_id)
            else:
                not_found_ids.append(request_id)

        conn.commit()
    finally:
        conn.close()

    if denied_ids:
        record_audit_event(
            "admin.credits.bulk_deny", request_id=request.state.request_id,
            actor_user_id=admin["id"], actor_username=admin["username"],
            target_type="credit_request",
            metadata={"denied_ids": denied_ids, "count": len(denied_ids)},
            ip_address=request.state.client_ip, user_agent=request.state.user_agent,
        )

    return {
        "denied": len(denied_ids),
        "skipped": len(skipped_ids),
        "not_found": len(not_found_ids),
        "request_ids": denied_ids,
    }


@app.post("/api/admin/users/{user_id}/credits")
async def admin_grant_credits(user_id: int, grant: AdminCreditGrant, request: Request, admin: dict = Depends(get_admin_user)):
    """Manually grant credits to a user (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    # Verify user exists
    cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    conn.close()

    new_balance = add_credits(user_id, grant.amount, grant.reason or "Admin manual grant", admin["id"])

    record_audit_event(
        "admin.credits.grant",
        request_id=getattr(request.state, "request_id", None),
        actor_user_id=admin["id"],
        actor_username=admin["username"],
        target_type="user",
        target_id=str(user_id),
        ip_address=getattr(request.state, "client_ip", None),
        user_agent=getattr(request.state, "user_agent", None),
        metadata={"amount": grant.amount, "reason": grant.reason},
    )

    return {"message": f"Granted {grant.amount} credits", "new_balance": new_balance}


@app.get("/api/admin/users/{user_id}/credits")
async def get_user_credits_admin(user_id: int, admin: dict = Depends(get_admin_user)):
    """Get a user's credit balance and history (admin only)"""
    conn = get_db()
    cursor = conn.cursor()

    # Get user info
    cursor.execute("SELECT username, email, credit_balance FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")

    # Get recent transactions
    cursor.execute("""
        SELECT id, job_id, amount, balance_after, transaction_type, reason, created_at
        FROM credit_ledger
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 20
    """, (user_id,))

    transactions = [
        {
            "id": row[0],
            "job_id": row[1],
            "amount": row[2],
            "balance_after": row[3],
            "transaction_type": row[4],
            "reason": row[5],
            "created_at": row[6]
        }
        for row in cursor.fetchall()
    ]
    conn.close()

    return {
        "user_id": user_id,
        "username": user[0],
        "email": user[1],
        "credit_balance": user[2],
        "transactions": transactions
    }

# ==================== SCRAPER ====================

class GoogleBusinessScraper:
    def __init__(self, headless: bool = None):
        """Create scraper instance. headless defaults to env PLAYWRIGHT_HEADLESS (True)"""
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        # Determine headless mode from param or environment
        if headless is None:
            self.headless = config.PLAYWRIGHT_HEADLESS
        else:
            self.headless = headless

    async def init_browser(self):
        """Initialize Playwright browser with proper async support for Windows"""
        try:
            logger.info('Initializing Playwright browser with async API')

            # Start Playwright
            self.playwright = await async_playwright().start()

            # Launch browser
            self.browser = await self.playwright.chromium.launch(
                headless=self.headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox'
                ]
            )

            # Create browser context with realistic settings
            user_agent = config.PLAYWRIGHT_USER_AGENT

            self.context = await self.browser.new_context(
                viewport={"width": 1280, "height": 800},
                user_agent=user_agent,
                locale="en-US",
                extra_http_headers={"accept-language": "en-US,en;q=0.9"}
            )

            # Create page
            self.page = await self.context.new_page()

            # Hide webdriver flag
            await self.page.add_init_script('Object.defineProperty(navigator, "webdriver", {get: () => undefined})')

            # Block images for faster loading
            await self.context.route("**/*.{png,jpg,jpeg,gif,webp,svg,ico}", lambda route: route.abort())

            logger.info('Playwright browser initialized successfully')

        except Exception as e:
            logger.exception(f'Failed to initialize Playwright browser: {e}')
            raise


    async def scrape_location(self, category: str, city: str, state: str, max_results: int = 10) -> List[Dict]:
        """Scrape businesses for a specific location using Playwright"""
        if not self.page:
            await self.init_browser()

        results = []
        search_term = f"{category} in {city}, {state}"

        try:
            logger.info(f"Starting scraping for {city}, {state}")

            search_url = f"https://www.google.com/maps/search/{quote_plus(search_term)}"
            logger.info(f"Searching Google Maps for: {search_term}")

            # Navigate to search page
            await self.page.goto(search_url, wait_until="networkidle", timeout=30000)

            # Wait for results to load
            try:
                await self.page.wait_for_selector('div[role="feed"]', timeout=10000)
                logger.info("Results feed loaded")
            except PlaywrightTimeout:
                logger.warning("Results feed did not load, trying alternative selector")
                await asyncio.sleep(2)

            # Scroll the results panel to load more businesses
            results_panel = await self.page.query_selector('div[role="feed"]')
            if results_panel:
                logger.info("Scrolling to load more results")
                for _ in range(3):
                    await results_panel.evaluate('el => el.scrollTop = el.scrollHeight')
                    await asyncio.sleep(1)

            # Extract business links - Google Maps uses <a> tags with specific href patterns
            business_links = await self.page.query_selector_all('a[href*="/maps/place/"]')
            logger.info(f"Found {len(business_links)} business links")

            # Get unique business URLs (avoid duplicates)
            business_urls = []
            seen_urls = set()
            for link in business_links[:max_results * 2]:  # Get extra in case of duplicates
                try:
                    href = await link.get_attribute('href')
                    if href and href not in seen_urls and '/maps/place/' in href:
                        business_urls.append(href)
                        seen_urls.add(href)
                        if len(business_urls) >= max_results:
                            break
                except Exception as e:
                    logger.debug(f"Error getting link href: {e}")
                    continue

            logger.info(f"Processing {len(business_urls)} unique businesses")

            # Visit each business page to get details
            for idx, url in enumerate(business_urls):
                try:
                    logger.info(f"Scraping business {idx + 1}/{len(business_urls)}")

                    # Navigate to business page
                    await self.page.goto(url, wait_until="domcontentloaded", timeout=15000)
                    await asyncio.sleep(1)

                    # Extract business name
                    business_name = "Unknown"
                    try:
                        name_elem = await self.page.query_selector('h1')
                        if name_elem:
                            business_name = (await name_elem.text_content()).strip()
                    except Exception:
                        pass

                    # Extract phone number
                    phone = ""
                    try:
                        phone_button = await self.page.query_selector('button[data-item-id*="phone"]')
                        if phone_button:
                            phone_text = await phone_button.get_attribute('aria-label')
                            if phone_text:
                                # Extract just the number from aria-label
                                import re
                                phone_match = re.search(r'[\d\s\-\(\)\+]+', phone_text)
                                if phone_match:
                                    phone = phone_match.group(0).strip()
                    except Exception:
                        pass

                    # Extract website
                    website = ""
                    try:
                        website_link = await self.page.query_selector('a[data-item-id*="authority"]')
                        if website_link:
                            website = await website_link.get_attribute('href')
                    except Exception:
                        pass

                    # Extract address
                    address = ""
                    try:
                        address_button = await self.page.query_selector('button[data-item-id*="address"]')
                        if address_button:
                            address_text = await address_button.get_attribute('aria-label')
                            if address_text:
                                # Remove "Address: " prefix if present
                                address = address_text.replace('Address:', '').strip()
                    except Exception:
                        pass

                    business_data = {
                        'business_name': business_name,
                        'phone': phone,
                        'website': website,
                        'address': address,
                        'category': category,
                        'city': city,
                        'state': state,
                        'google_maps_url': url
                    }

                    results.append(business_data)
                    logger.info(f"Scraped: {business_name}")

                except PlaywrightTimeout:
                    logger.warning(f"Timeout loading business page: {url}")
                    continue
                except Exception as e:
                    logger.debug(f"Error scraping business at {url}: {e}")
                    continue

            logger.info(f"Found {len(results)} businesses for {city}, {state}")

        except Exception as e:
            logger.exception(f"Error scraping location {city}, {state}: {e}")
            return []

        return results

    async def close(self):
        """Close browser and Playwright safely"""
        try:
            if self.page:
                await self.page.close()
            if self.context:
                await self.context.close()
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
            logger.info("Browser closed successfully")
        except Exception as e:
            logger.debug(f"Error closing browser: {e}")

async def run_scraping_job(job_id: str, request: ScrapingRequest, user_id: int):
    """Background task for running scraping jobs"""
    scraper = GoogleBusinessScraper()
    conn = get_db()
    
    try:
        logger.info(f"Starting scraping job {job_id}")
        # Debug: log the incoming list of city/state strings
        logger.debug(f"Received {len(request.cities_data)} city/state entries: {request.cities_data}")
        
        # Update job status
        cursor = conn.cursor()
        cursor.execute("UPDATE jobs SET status = ? WHERE job_id = ?", ("running", job_id))
        conn.commit()
        
        # Add log
        cursor.execute("INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                      (job_id, f"Starting scraping job for category: {request.category}", datetime.now().isoformat()))
        conn.commit()
        
        all_results = []
        
        for idx, city_state in enumerate(request.cities_data):
            # Check if job was cancelled
            cursor.execute("SELECT status FROM jobs WHERE job_id = ?", (job_id,))
            job_status = cursor.fetchone()
            if job_status and job_status[0] in ("failed", "cancelled"):
                break
            
            # Parse city and state
            # Log the raw entry for debugging
            logger.debug(f"Raw city_state entry: '{city_state}'")
            parts = city_state.split(",")
            if len(parts) >= 2:
                city = parts[0].strip()
                state = parts[1].strip()
                logger.debug(f"Parsed city: '{city}', state: '{state}'")
                
                current_city = f"{city}, {state}"
                progress = int((idx + 1) / len(request.cities_data) * 100)
                
                # Update job progress
                cursor.execute("""
                    UPDATE jobs SET current_city = ?, progress = ?
                    WHERE job_id = ?
                """, (current_city, progress, job_id))
                conn.commit()
                
                # Add log
                cursor.execute("INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                              (job_id, f"Processing city {idx + 1}/{len(request.cities_data)}: {city}, {state}", datetime.now().isoformat()))
                conn.commit()
                
                logger.info(f"Processing city {idx + 1}/{len(request.cities_data)}: {city}, {state}")
                
                # Scrape this location
                results = await scraper.scrape_location(
                    request.category, 
                    city, 
                    state, 
                    request.max_results_per_city
                )
                
                all_results.extend(results)
                
                # Save results to database
                for result in results:
                    cursor.execute("""
                        INSERT INTO results (job_id, business_name, phone, website, address, category, city, state, google_maps_url)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        job_id, result['business_name'], result['phone'], result['website'],
                        result['address'], result['category'], result['city'], result['state'],
                        result['google_maps_url']
                    ))
                conn.commit()
                
                # Add log
                cursor.execute("INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                              (job_id, f"Found {len(results)} businesses in {city}, {state}", datetime.now().isoformat()))
                conn.commit()
                
                logger.info(f"Found {len(results)} businesses in {city}, {state}")
                
                await asyncio.sleep(1)
        
        # If job was cancelled/failed, do not mark it as completed.
        cursor.execute("SELECT status FROM jobs WHERE job_id = ?", (job_id,))
        final_status_row = cursor.fetchone()
        if final_status_row and final_status_row[0] in ("failed", "cancelled"):
            if final_status_row[0] == "cancelled":
                cursor.execute(
                    "INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                    (job_id, "Job cancelled by user", datetime.now().isoformat()),
                )
                conn.commit()
            return

        # Save results to CSV file
        if all_results:
            filename = f"{RESULTS_DIR}/results_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            df = pd.DataFrame(all_results)
            df.to_csv(filename, index=False)
            logger.info(f"Saved {len(all_results)} results to {filename}")
        
        # Update job as completed
        cursor.execute("""
            UPDATE jobs SET status = ?, progress = ?, completed_at = ?
            WHERE job_id = ?
        """, ("completed", 100, datetime.now().isoformat(), job_id))
        conn.commit()
        
        # Add completion log
        cursor.execute("INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                      (job_id, f"Job completed successfully. Total businesses found: {len(all_results)}", datetime.now().isoformat()))
        conn.commit()
        
        logger.info(f"Job {job_id} completed successfully. Total businesses found: {len(all_results)}")
        
    except Exception as e:
        # Log full traceback for diagnostics
        logger.exception(f"Job {job_id} failed with error: {repr(e)}")
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE jobs SET status = ?, error = ?, completed_at = ?
            WHERE job_id = ?
        """, ("failed", repr(e), datetime.now().isoformat(), job_id))
        conn.commit()
        
        cursor.execute("INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
                      (job_id, f"Job failed with error: {repr(e)}", datetime.now().isoformat()))
        conn.commit()
    
    finally:
        conn.close()
        await scraper.close()
        logger.info(f"Closed browser for job {job_id}")

# ==================== JOB ENDPOINTS ====================

@app.post("/api/jobs", response_model=ScrapingJob)
async def create_scraping_job(
    job_request: ScrapingRequest,
    req: Request,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Create a new scraping job with credit check and rate limiting"""
    user_id = current_user["id"]

    # Check rate limits
    rate_status = check_rate_limits(user_id)
    if not rate_status["can_proceed"]:
        raise HTTPException(status_code=429, detail=rate_status["reason"])

    # Calculate credit cost
    num_cities = len(job_request.cities_data)
    credit_cost = estimate_credit_cost(num_cities, job_request.max_results_per_city)
    current_balance = get_user_credit_balance(user_id)

    # Check sufficient credits
    if current_balance < credit_cost:
        raise HTTPException(
            status_code=402,  # Payment Required
            detail=f"Insufficient credits. Required: {credit_cost}, Available: {current_balance}"
        )

    # Deduct credits upfront
    new_balance = deduct_credits(user_id, credit_cost, reason=f"Job charge: {job_request.category}")

    job_id = str(uuid.uuid4())

    conn = get_db()
    cursor = conn.cursor()

    # Create job in database with credit info
    cursor.execute("""
        INSERT INTO jobs (job_id, user_id, category, cities_data, max_results_per_city, status, total_cities, credit_estimate, credit_charged, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        job_id, user_id, job_request.category,
        json.dumps(job_request.cities_data), job_request.max_results_per_city,
        "pending", num_cities, credit_cost, credit_cost, datetime.now().isoformat()
    ))
    conn.commit()

    # Update the credit ledger with job_id
    cursor.execute("""
        UPDATE credit_ledger SET job_id = ?
        WHERE user_id = ? AND job_id IS NULL
        ORDER BY created_at DESC LIMIT 1
    """, (job_id, user_id))
    conn.commit()
    conn.close()

    # Increment rate limit counter
    increment_rate_limit(user_id)

    # Start background task
    background_tasks.add_task(run_scraping_job, job_id, job_request, user_id)

    # Audit: job create
    record_audit_event(
        "job.create",
        request_id=req.state.request_id,
        actor_user_id=current_user["id"],
        actor_username=current_user["username"],
        target_type="job",
        target_id=job_id,
        after={"category": job_request.category, "cities": num_cities, "credit_cost": credit_cost},
        ip_address=req.state.client_ip,
        user_agent=req.state.user_agent,
    )

    # Return job
    return ScrapingJob(
        job_id=job_id,
        status=ScrapingStatus.PENDING,
        created_at=datetime.now().isoformat(),
        logs=[f"Job created for category: {job_request.category}", f"Credits charged: {credit_cost}"]
    )


@app.post("/api/jobs/{job_id}/cancel")
async def cancel_scraping_job(
    job_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """Cancel a pending/running scraping job (best-effort)."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT user_id, status FROM jobs WHERE job_id = ?", (job_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")

    job_user_id, status_value = row[0], row[1]
    if job_user_id != current_user["id"] and current_user.get("is_admin") is not True:
        conn.close()
        raise HTTPException(status_code=403, detail="Not authorized to cancel this job")

    if status_value in ("completed", "failed", "cancelled"):
        conn.close()
        return {"message": f"Job already {status_value}"}

    now = datetime.now().isoformat()
    cursor.execute(
        """
        UPDATE jobs
        SET status = ?, error = ?, completed_at = ?
        WHERE job_id = ?
        """,
        ("cancelled", "Cancelled by user", now, job_id),
    )
    cursor.execute(
        "INSERT INTO job_logs (job_id, log_message, created_at) VALUES (?, ?, ?)",
        (job_id, "Cancellation requested by user", now),
    )

    conn.commit()
    conn.close()

    # Audit: job cancel
    record_audit_event(
        "job.cancel",
        request_id=request.state.request_id,
        actor_user_id=current_user["id"],
        actor_username=current_user["username"],
        target_type="job",
        target_id=job_id,
        before={"status": status_value},
        after={"status": "cancelled"},
        ip_address=request.state.client_ip,
        user_agent=request.state.user_agent,
    )

    return {"message": "Cancellation requested"}


@app.delete("/api/jobs/{job_id}")
async def delete_scraping_job(
    job_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """Delete a completed/failed/cancelled job and its results/logs (owner or admin)."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT user_id, status, category FROM jobs WHERE job_id = ?", (job_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")

    job_user_id, status_value, category = row[0], row[1], row[2]
    if job_user_id != current_user["id"] and current_user.get("is_admin") is not True:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")

    if status_value in ("pending", "running"):
        conn.close()
        raise HTTPException(
            status_code=409,
            detail="Job is running. Cancel it first, then delete.",
        )

    cursor.execute("DELETE FROM jobs WHERE job_id = ?", (job_id,))
    conn.commit()
    conn.close()

    # Audit: job delete
    record_audit_event(
        "job.delete",
        request_id=request.state.request_id,
        actor_user_id=current_user["id"],
        actor_username=current_user["username"],
        target_type="job",
        target_id=job_id,
        before={"status": status_value, "category": category},
        ip_address=request.state.client_ip,
        user_agent=request.state.user_agent,
    )

    return {"message": "Job deleted"}

@app.get("/api/jobs/{job_id}", response_model=ScrapingJob)
async def get_job_status(job_id: str, current_user: dict = Depends(get_current_user)):
    """Get scraping job status"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get job
    cursor.execute("""
        SELECT job_id, status, progress, total_cities, current_city, error, created_at, completed_at
        FROM jobs WHERE job_id = ? AND user_id = ?
    """, (job_id, current_user["id"]))
    
    job_data = cursor.fetchone()
    if not job_data:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get logs
    cursor.execute("SELECT log_message FROM job_logs WHERE job_id = ? ORDER BY created_at", (job_id,))
    logs = [row[0] for row in cursor.fetchall()]
    
    # Get results count
    cursor.execute("SELECT COUNT(*) FROM results WHERE job_id = ?", (job_id,))
    results_count = cursor.fetchone()[0]
    
    # Get results
    cursor.execute("""
        SELECT business_name, phone, website, address, category, city, state, google_maps_url
        FROM results WHERE job_id = ?
    """, (job_id,))
    
    results = []
    for row in cursor.fetchall():
        results.append({
            "business_name": row[0],
            "phone": row[1],
            "website": row[2],
            "address": row[3],
            "category": row[4],
            "city": row[5],
            "state": row[6],
            "google_maps_url": row[7]
        })
    
    conn.close()
    
    return ScrapingJob(
        job_id=job_data[0],
        status=ScrapingStatus(job_data[1]),
        progress=job_data[2],
        total_cities=job_data[3],
        current_city=job_data[4] or "",
        results=results,
        error=job_data[5],
        created_at=job_data[6],
        completed_at=job_data[7],
        logs=logs
    )

@app.get("/api/jobs")
async def get_user_jobs(current_user: dict = Depends(get_current_user)):
    """Get all jobs for current user"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT job_id, category, status, progress, created_at, completed_at
        FROM jobs WHERE user_id = ? ORDER BY created_at DESC
    """, (current_user["id"],))
    
    jobs = []
    for row in cursor.fetchall():
        jobs.append({
            "job_id": row[0],
            "category": row[1],
            "status": row[2],
            "progress": row[3],
            "created_at": row[4],
            "completed_at": row[5]
        })
    
    conn.close()
    return {"jobs": jobs}

@app.get("/api/jobs/{job_id}/results")
async def get_job_results(job_id: str, current_user: dict = Depends(get_current_user)):
    """Get scraping results"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Verify job ownership
    cursor.execute("SELECT user_id FROM jobs WHERE job_id = ?", (job_id,))
    job = cursor.fetchone()
    if not job or job[0] != current_user["id"]:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get results
    cursor.execute("""
        SELECT business_name, phone, website, address, category, city, state, google_maps_url
        FROM results WHERE job_id = ?
    """, (job_id,))
    
    results = []
    for row in cursor.fetchall():
        results.append({
            "business_name": row[0],
            "phone": row[1],
            "website": row[2],
            "address": row[3],
            "category": row[4],
            "city": row[5],
            "state": row[6],
            "google_maps_url": row[7]
        })
    
    conn.close()
    
    return {
        "job_id": job_id,
        "total_results": len(results),
        "results": results
    }

def create_formatted_xlsx(results: List[Dict], category: str, location: str) -> bytes:
    """Create a professionally formatted Excel file"""
    wb = Workbook()
    ws = wb.active

    # Generate sheet title from category and location
    sheet_title = f"{location}_{category}".replace(" ", "_")[:31]  # Excel sheet name limit
    ws.title = sheet_title

    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Calibri', size=11)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    link_font = Font(name='Calibri', size=11, color='0563C1', underline='single')

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Define column headers and widths
    headers = ['#', 'Business Name', 'Phone', 'Website', 'Address', 'City', 'State', 'Google Maps']
    col_widths = [5, 35, 18, 40, 50, 20, 15, 45]

    # Write headers
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Write data
    for row_idx, result in enumerate(results, 2):
        # Row number
        cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # Business Name
        cell = ws.cell(row=row_idx, column=2, value=result.get('business_name', 'N/A'))
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Phone
        phone = result.get('phone', '') or 'N/A'
        cell = ws.cell(row=row_idx, column=3, value=phone)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Website (as hyperlink if valid)
        website = result.get('website', '') or 'N/A'
        cell = ws.cell(row=row_idx, column=4)
        if website and website != 'N/A' and website.startswith('http'):
            cell.value = website
            cell.hyperlink = website
            cell.font = link_font
        else:
            cell.value = website
            cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Address
        address = result.get('address', '') or 'N/A'
        cell = ws.cell(row=row_idx, column=5, value=address)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # City
        cell = ws.cell(row=row_idx, column=6, value=result.get('city', 'N/A'))
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # State
        cell = ws.cell(row=row_idx, column=7, value=result.get('state', 'N/A'))
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Google Maps URL (as hyperlink)
        maps_url = result.get('google_maps_url', '') or 'N/A'
        cell = ws.cell(row=row_idx, column=8)
        if maps_url and maps_url != 'N/A' and maps_url.startswith('http'):
            cell.value = 'View on Maps'
            cell.hyperlink = maps_url
            cell.font = link_font
        else:
            cell.value = 'N/A'
            cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # Alternate row coloring
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Add auto-filter
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@app.get("/api/jobs/{job_id}/download")
async def download_results(job_id: str, current_user: dict = Depends(get_current_user)):
    """Download results as formatted Excel (.xlsx) file"""
    conn = get_db()
    cursor = conn.cursor()

    # Verify job ownership and get job details
    cursor.execute("SELECT status, category FROM jobs WHERE job_id = ? AND user_id = ?", (job_id, current_user["id"]))
    job = cursor.fetchone()
    if not job:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")

    if job[0] != "completed":
        conn.close()
        raise HTTPException(status_code=400, detail="Job not completed")

    category = job[1] or "Business"

    # Get results
    cursor.execute("""
        SELECT business_name, phone, website, address, category, city, state, google_maps_url
        FROM results WHERE job_id = ?
    """, (job_id,))

    results = []
    cities_set = set()
    for row in cursor.fetchall():
        results.append({
            "business_name": row[0],
            "phone": row[1],
            "website": row[2],
            "address": row[3],
            "category": row[4],
            "city": row[5],
            "state": row[6],
            "google_maps_url": row[7]
        })
        if row[5]:
            cities_set.add(row[5])

    conn.close()

    # Generate location name for filename
    if len(cities_set) == 1:
        location = list(cities_set)[0].replace(" ", "_")
    elif len(cities_set) > 1:
        location = f"{len(cities_set)}_Cities"
    else:
        location = "Results"

    # Create formatted Excel file
    xlsx_bytes = create_formatted_xlsx(results, category, location)
    xlsx_base64 = base64.b64encode(xlsx_bytes).decode('utf-8')

    # Generate filename: Location_Category.xlsx (e.g., "City_of_London_Hotels.xlsx")
    safe_category = category.replace(" ", "_").replace("/", "_")
    safe_location = location.replace(" ", "_").replace("/", "_")
    filename = f"{safe_location}_{safe_category}.xlsx"

    return {
        "filename": filename,
        "content": xlsx_base64,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "encoding": "base64"
    }

# ==================== PAYMENTS & SUBSCRIPTIONS ====================

_payment_processor = PaymentProcessor()
_pricing_engine = PricingEngine()
_promo_service = PromoService()
_subscription_manager = SubscriptionManager()
_webhook_handler = WebhookHandler()
_coinbase_service = CoinbaseCommerceService()


@app.get("/api/payments/packages")
async def get_credit_packages(current_user: dict = Depends(get_current_user)):
    """List active credit packages"""
    from db.payment_crud import list_active_packages

    packages = []
    for row in list_active_packages():
        # credit_packages schema: (id, name, credits, base_price_cents, display_price_cents, discount_percentage,
        #                          tier_id, is_active, is_featured, features, created_at, updated_at)
        features = None
        try:
            features = json.loads(row[9]) if row[9] else None
        except Exception:
            features = None

        packages.append({
            "id": row[0],
            "name": row[1],
            "credits": row[2],
            "base_price_cents": row[3],
            "display_price_cents": row[4],
            "discount_percentage": row[5],
            "tier_id": row[6],
            "is_active": bool(row[7]),
            "is_featured": bool(row[8]),
            "features": features,
        })

    return packages


@app.get("/api/payments/packages/{package_id}")
async def get_credit_package_details(package_id: int, current_user: dict = Depends(get_current_user)):
    """Get a single credit package"""
    from db.payment_crud import get_credit_package

    row = get_credit_package(package_id)
    if not row or not bool(row[7]):
        raise HTTPException(status_code=404, detail="Package not found")

    features = None
    try:
        features = json.loads(row[9]) if row[9] else None
    except Exception:
        features = None

    return {
        "id": row[0],
        "name": row[1],
        "credits": row[2],
        "base_price_cents": row[3],
        "display_price_cents": row[4],
        "discount_percentage": row[5],
        "tier_id": row[6],
        "is_active": bool(row[7]),
        "is_featured": bool(row[8]),
        "features": features,
    }


@app.post("/api/payments/calculate-price")
async def calculate_price(req: PriceCalcRequest, current_user: dict = Depends(get_current_user)):
    """Calculate final price for a credit package (tier + promo + tax)"""
    try:
        price = _pricing_engine.calculate_for_package(
            user_id=current_user["id"],
            package_id=req.package_id,
            quantity=req.quantity,
            promo_code=req.promo_code,
        )
        return {
            "subtotal_cents": price["subtotal_cents"],
            "discount_cents": price["discount_cents"],
            "tax_cents": price["tax_cents"],
            "total_cents": price["total_cents"],
            "credits_to_receive": price["credits_to_receive"],
            "bonus_credits": price["bonus_credits"],
            "promo_code_applied": price["promo"].code if price.get("promo") else None,
            "pricing_tier": price["tier"].name if price.get("tier") else None,
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/payments/purchase")
async def purchase_credits(req: PurchaseRequest, current_user: dict = Depends(get_current_user)):
    """Initiate a Stripe PaymentIntent or Coinbase charge for a credit package"""
    if req.quantity <= 0:
        raise HTTPException(status_code=400, detail="quantity must be > 0")

    provider = (req.provider or "stripe").lower()
    if provider not in ("stripe", "coinbase"):
        raise HTTPException(status_code=400, detail="provider must be 'stripe' or 'coinbase'")

    try:
        if provider == "stripe":
            result = _payment_processor.initiate_stripe_package_purchase(
                user_id=current_user["id"],
                email=current_user["email"],
                package_id=req.package_id,
                quantity=req.quantity,
                promo_code=req.promo_code,
                idempotency_key=req.idempotency_key,
                currency=config.PAYMENT_CONFIG.get("currency", "USD"),
            )
            return {
                "transaction_id": result.transaction_id,
                "payment_provider": result.provider,
                "amount_cents": result.amount_cents,
                "currency": result.currency,
                "credits_purchased": result.credits,
                "bonus_credits": result.bonus_credits,
                "provider_transaction_id": result.provider_transaction_id,
                "client_secret": result.client_secret,
            }

        # coinbase
        result = _payment_processor.initiate_coinbase_package_purchase(
            user_id=current_user["id"],
            package_id=req.package_id,
            quantity=req.quantity,
            promo_code=req.promo_code,
            idempotency_key=req.idempotency_key,
            currency=config.PAYMENT_CONFIG.get("currency", "USD"),
        )
        return {
            "transaction_id": result.transaction_id,
            "payment_provider": result.provider,
            "amount_cents": result.amount_cents,
            "currency": result.currency,
            "credits_purchased": result.credits,
            "bonus_credits": result.bonus_credits,
            "provider_transaction_id": result.provider_transaction_id,
            "hosted_url": result.hosted_url,
        }

    except ValueError as e:
        # Idempotency conflict etc
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class SimplePurchaseRequest(BaseModel):
    credits: int  # Must be >= 100
    promo_code: Optional[str] = None


@app.post("/api/payments/simple-purchase")
async def simple_purchase(
    request: SimplePurchaseRequest,
    current_user: dict = Depends(get_current_user)
):
    """Simple credit purchase: $0.10 per credit (flat rate)"""
    if request.credits < 100:
        raise HTTPException(status_code=400, detail="Minimum purchase is 100 credits")

    # Calculate price in cents
    base_cents = request.credits * 10  # $0.10 = 10 cents per credit
    discount_cents = 0
    promo_name = ""
    bonus_credits = 0

    # Apply promo code if provided
    if request.promo_code:
        try:
            normalized_code = str(request.promo_code).strip().upper()
            promo_effect = _promo_service.validate_and_calculate(
                user_id=current_user["id"],
                code=normalized_code,
                subtotal_cents=int(base_cents),
                applies_to="all",
            )
            if promo_effect:
                discount_cents = int(promo_effect.discount_cents or 0)
                bonus_credits = int(promo_effect.bonus_credits or 0)
                promo_name = str(promo_effect.code or normalized_code)
                logging.info(
                    "Applied promo on simple-purchase user_id=%s code=%s discount_cents=%s bonus_credits=%s",
                    current_user["id"],
                    promo_name,
                    discount_cents,
                    bonus_credits,
                )
        except Exception as e:
            logging.warning(f"Promo code validation failed: {e}")

    total_cents = max(0, base_cents - discount_cents)

    # Create Stripe checkout session
    try:
        import stripe
        stripe.api_key = config.STRIPE_SECRET_KEY

        session = stripe.checkout.Session.create(
            line_items=[{
                'price_data': {
                    'currency': 'usd',
                    'unit_amount': total_cents,
                    'product_data': {'name': f'{request.credits} Credits'},
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url=f'{config.FRONTEND_URL}/payment/success?session_id={{CHECKOUT_SESSION_ID}}',
            cancel_url=f'{config.FRONTEND_URL}/wallet',
            metadata={
                'user_id': str(current_user['id']),
                'credits': str(request.credits),
                'promo_code': (promo_name or (request.promo_code or "")).strip(),
                'subtotal_cents': str(int(base_cents)),
                'discount_cents': str(int(discount_cents)),
                'bonus_credits': str(int(bonus_credits)),
            }
        )

        return {
            "checkout_url": session.url,
            "session_id": session.id,
            "credits": request.credits,
            "subtotal_cents": base_cents,
            "discount_cents": discount_cents,
            "total_cents": total_cents,
            "bonus_credits": bonus_credits,
            "promo_applied": promo_name if promo_name else None,
        }
    except Exception as e:
        logging.error(f"Stripe session creation failed: {e}")
        raise HTTPException(status_code=400, detail="Failed to create checkout session")


@app.post("/api/payments/crypto/create-charge")
async def create_crypto_charge(req: PurchaseRequest, current_user: dict = Depends(get_current_user)):
    """
    Back-compat alias for initiating a Coinbase Commerce purchase.
    Prefer POST /api/payments/purchase with provider='coinbase'.
    """
    req.provider = "coinbase"
    return await purchase_credits(req, current_user=current_user)


@app.get("/api/payments/methods")
async def list_payment_methods(current_user: dict = Depends(get_current_user)):
    """List saved payment methods"""
    from db.payment_crud import get_user_payment_methods

    rows = get_user_payment_methods(current_user["id"])
    methods = []
    for r in rows:
        methods.append({
            "id": r[0],
            "user_id": r[1],
            "stripe_payment_method_id": r[2],
            "type": r[3],
            "is_default": bool(r[4]),
            "card_brand": r[5],
            "card_last4": r[6],
            "card_exp_month": r[7],
            "card_exp_year": r[8],
            "billing_name": r[9],
            "billing_email": r[10],
            "billing_address": r[11],
            "created_at": r[12],
            "updated_at": r[13],
        })
    return methods


@app.post("/api/payments/methods")
async def add_payment_method(req: AddPaymentMethodRequest, current_user: dict = Depends(get_current_user)):
    """Store a Stripe payment method ID and (best-effort) attach it to the user's Stripe customer."""
    from db.payment_crud import create_payment_method, set_default_payment_method

    # Best-effort Stripe attach + card details fetch
    card_brand = None
    card_last4 = None
    card_exp_month = None
    card_exp_year = None

    try:
        import stripe
        # If keys are left as placeholders, skip talking to Stripe.
        if config.STRIPE_SECRET_KEY and not config.STRIPE_SECRET_KEY.startswith("sk_test_..."):
            stripe.api_key = config.STRIPE_SECRET_KEY
            # Ensure customer exists
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("SELECT stripe_customer_id FROM users WHERE id = ?", (current_user["id"],))
            customer_id = cursor.fetchone()[0]
            conn.close()

            if not customer_id:
                customer = stripe.Customer.create(email=current_user["email"], metadata={"user_id": str(current_user["id"])})
                customer_id = customer["id"]
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET stripe_customer_id = ? WHERE id = ?", (customer_id, current_user["id"]))
                conn.commit()
                conn.close()

            stripe.PaymentMethod.attach(req.stripe_payment_method_id, customer=str(customer_id))
            pm = stripe.PaymentMethod.retrieve(req.stripe_payment_method_id)
            card = pm.get("card") or {}
            card_brand = card.get("brand")
            card_last4 = card.get("last4")
            card_exp_month = card.get("exp_month")
            card_exp_year = card.get("exp_year")
    except Exception:
        # Don't block storing the ID if Stripe is unreachable/misconfigured during dev.
        pass

    method_id = create_payment_method(
        user_id=current_user["id"],
        stripe_payment_method_id=req.stripe_payment_method_id,
        method_type="card",
        is_default=req.is_default,
        card_brand=card_brand,
        card_last4=card_last4,
        card_exp_month=card_exp_month,
        card_exp_year=card_exp_year,
    )

    if req.is_default:
        set_default_payment_method(current_user["id"], int(method_id))

    return {"id": int(method_id)}


@app.delete("/api/payments/methods/{method_id}")
async def remove_payment_method(method_id: int, current_user: dict = Depends(get_current_user)):
    from db.payment_crud import delete_payment_method

    # Minimal safety: delete by ID (CRUD helper doesn't verify user_id)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT user_id, stripe_payment_method_id FROM payment_methods WHERE id = ?", (method_id,))
    row = cursor.fetchone()
    conn.close()
    if not row or row[0] != current_user["id"]:
        raise HTTPException(status_code=404, detail="Payment method not found")

    stripe_pm_id = row[1]
    try:
        import stripe
        if config.STRIPE_SECRET_KEY and not config.STRIPE_SECRET_KEY.startswith("sk_test_..."):
            stripe.api_key = config.STRIPE_SECRET_KEY
            stripe.PaymentMethod.detach(stripe_pm_id)
    except Exception:
        pass

    ok = delete_payment_method(method_id)
    return {"deleted": bool(ok)}


@app.put("/api/payments/methods/{method_id}/default")
async def set_default_method(method_id: int, current_user: dict = Depends(get_current_user)):
    from db.payment_crud import set_default_payment_method

    ok = set_default_payment_method(current_user["id"], method_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Payment method not found")
    return {"ok": True}


@app.get("/api/payments/transactions")
async def list_my_transactions(current_user: dict = Depends(get_current_user)):
    """List current user's payment transactions"""
    from db.payment_crud import get_user_transactions

    rows = get_user_transactions(current_user["id"], limit=100)
    txns = []
    for r in rows:
        txns.append({
            "id": r[0],
            "transaction_id": r[1],
            "user_id": r[2],
            "payment_provider": r[3],
            "provider_transaction_id": r[4],
            "amount_cents": r[5],
            "currency": r[6],
            "status": r[7],
            "credits_purchased": r[8],
            "idempotency_key": r[9],
            "package_id": r[10],
            "subscription_id": r[11],
            "promo_code_id": r[12],
            "invoice_id": r[13],
            "created_at": r[14],
            "updated_at": r[15],
        })
    return {"transactions": txns}


@app.get("/api/payments/transactions/{transaction_id}")
async def get_transaction_details(transaction_id: str, current_user: dict = Depends(get_current_user)):
    """Get payment transaction details (and provider status if available)"""
    from db.payment_crud import get_transaction

    txn = get_transaction(transaction_id)
    if not txn or txn[2] != current_user["id"]:
        raise HTTPException(status_code=404, detail="Transaction not found")

    resp = {
        "id": txn[0],
        "transaction_id": txn[1],
        "user_id": txn[2],
        "payment_provider": txn[3],
        "provider_transaction_id": txn[4],
        "amount_cents": txn[5],
        "currency": txn[6],
        "status": txn[7],
        "credits_purchased": txn[8],
        "invoice_id": txn[13],
        "created_at": txn[14],
        "updated_at": txn[15],
    }

    # Optional: enrich Coinbase status
    if resp["payment_provider"] == "coinbase" and resp.get("provider_transaction_id"):
        try:
            charge = _coinbase_service.get_charge(resp["provider_transaction_id"])
            timeline = charge.get("timeline") or []
            last_status = timeline[-1]["status"] if timeline else None
            resp["provider_status"] = last_status
        except Exception:
            resp["provider_status"] = None

    return resp


@app.get("/api/payments/crypto/status/{charge_id}")
async def get_crypto_status(charge_id: str, current_user: dict = Depends(get_current_user)):
    """Poll Coinbase Commerce charge status"""
    try:
        charge = _coinbase_service.get_charge(charge_id)
        timeline = charge.get("timeline") or []
        status_value = timeline[-1]["status"] if timeline else "NEW"
        return {"charge_id": charge_id, "status": status_value, "hosted_url": charge.get("hosted_url")}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/webhooks/stripe")
async def stripe_webhook(request: Request):
    """Stripe webhook endpoint (signature-verified)"""
    payload = await request.body()
    sig = request.headers.get("stripe-signature")
    if not sig:
        raise HTTPException(status_code=400, detail="Missing Stripe signature header")

    try:
        return _webhook_handler.handle_stripe_webhook(payload, sig)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/webhooks/coinbase")
async def coinbase_webhook(request: Request):
    """Coinbase Commerce webhook endpoint (signature-verified)"""
    payload = await request.body()
    sig = request.headers.get("x-cc-webhook-signature") or request.headers.get("X-CC-Webhook-Signature")
    if not sig:
        raise HTTPException(status_code=400, detail="Missing Coinbase signature header")

    try:
        return _webhook_handler.handle_coinbase_webhook(payload, sig)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/subscriptions/plans")
async def list_subscription_plans(current_user: dict = Depends(get_current_user)):
    """List active subscription plans"""
    from db.payment_crud import list_active_subscription_plans

    plans = []
    for row in list_active_subscription_plans():
        # schema: id, name, billing_interval, credits_per_period, base_price_cents, stripe_price_id,
        # rollover_credits, max_rollover_credits, trial_days, is_active, is_featured, features, created_at, updated_at
        features = None
        try:
            features = json.loads(row[11]) if row[11] else None
        except Exception:
            features = None
        plans.append({
            "id": row[0],
            "name": row[1],
            "billing_interval": row[2],
            "credits_per_period": row[3],
            "base_price_cents": row[4],
            "stripe_price_id": row[5],
            "rollover_credits": bool(row[6]),
            "max_rollover_credits": row[7],
            "trial_days": row[8],
            "is_active": bool(row[9]),
            "is_featured": bool(row[10]),
            "features": features,
        })
    return plans


# ==================== INVOICES ====================

@app.get("/api/invoices")
async def list_invoices(current_user: dict = Depends(get_current_user)):
    from db.payment_crud import get_user_invoices

    rows = get_user_invoices(current_user["id"])
    invoices = []
    for r in rows:
        invoices.append({
            "id": r[0],
            "invoice_number": r[1],
            "user_id": r[2],
            "transaction_id": r[3],
            "subscription_id": r[4],
            "invoice_type": r[5],
            "amount_cents": r[6],
            "tax_amount_cents": r[7],
            "total_amount_cents": r[8],
            "status": r[9],
            "pdf_path": r[10],
            "created_at": r[15],
        })
    return {"invoices": invoices}


@app.get("/api/invoices/{invoice_id}")
async def invoice_details(invoice_id: int, current_user: dict = Depends(get_current_user)):
    from db.payment_crud import get_invoice_by_id

    inv = get_invoice_by_id(invoice_id)
    if not inv or inv[2] != current_user["id"]:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {
        "id": inv[0],
        "invoice_number": inv[1],
        "user_id": inv[2],
        "transaction_id": inv[3],
        "subscription_id": inv[4],
        "invoice_type": inv[5],
        "amount_cents": inv[6],
        "tax_amount_cents": inv[7],
        "total_amount_cents": inv[8],
        "status": inv[9],
        "pdf_path": inv[10],
        "created_at": inv[15],
        "updated_at": inv[16],
    }


@app.get("/api/invoices/{invoice_id}/download")
async def download_invoice(invoice_id: int, current_user: dict = Depends(get_current_user)):
    from db.payment_crud import get_invoice_by_id
    from invoices.invoice_service import ensure_invoice_pdf

    inv = get_invoice_by_id(invoice_id)
    if not inv or inv[2] != current_user["id"]:
        raise HTTPException(status_code=404, detail="Invoice not found")

    _, pdf_bytes = ensure_invoice_pdf(invoice_id)
    b64 = base64.b64encode(pdf_bytes).decode("utf-8")
    filename = f"{inv[1]}.pdf"
    return {
        "filename": filename,
        "content": b64,
        "content_type": "application/pdf",
        "encoding": "base64",
    }


@app.post("/api/subscriptions/subscribe")
async def subscribe(req: SubscribeRequest, current_user: dict = Depends(get_current_user)):
    """Create a Stripe subscription and return the payment client_secret if needed"""
    try:
        result = _subscription_manager.create_subscription(
            user_id=current_user["id"],
            email=current_user["email"],
            plan_id=req.plan_id,
            idempotency_key=req.idempotency_key,
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/subscriptions/my-subscription")
async def my_subscription(current_user: dict = Depends(get_current_user)):
    """Get current user's active subscription record"""
    from db.payment_crud import get_user_active_subscription

    sub = get_user_active_subscription(current_user["id"])
    if not sub:
        raise HTTPException(status_code=404, detail="No active subscription")
    return {
        "id": sub[0],
        "user_id": sub[1],
        "subscription_plan_id": sub[2],
        "stripe_subscription_id": sub[3],
        "status": sub[4],
        "current_period_start": sub[5],
        "current_period_end": sub[6],
        "next_billing_date": sub[7],
        "cancel_at_period_end": bool(sub[8]) if len(sub) > 8 else False,
    }


@app.put("/api/subscriptions/{stripe_subscription_id}/cancel")
async def cancel_subscription(stripe_subscription_id: str, immediate: bool = False, current_user: dict = Depends(get_current_user)):
    """Cancel subscription (at period end by default)"""
    try:
        return _subscription_manager.cancel(stripe_subscription_id, at_period_end=not immediate)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/subscriptions/{stripe_subscription_id}/pause")
async def pause_subscription(stripe_subscription_id: str, current_user: dict = Depends(get_current_user)):
    try:
        return _subscription_manager.pause(stripe_subscription_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/subscriptions/{stripe_subscription_id}/resume")
async def resume_subscription(stripe_subscription_id: str, current_user: dict = Depends(get_current_user)):
    try:
        return _subscription_manager.resume(stripe_subscription_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/promos/validate")
async def validate_promo(req: PromoValidateRequest, current_user: dict = Depends(get_current_user)):
    """Validate a promo code and return its calculated effect for the provided context."""
    code = (req.code or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="code is required")

    applies_to = (req.applies_to or "packages").strip().lower()
    if applies_to not in ("packages", "all"):
        raise HTTPException(status_code=400, detail="applies_to must be 'packages' or 'all'")

    # Compute a baseline subtotal without promo, then validate against it.
    if applies_to == "packages":
        if not req.package_id:
            raise HTTPException(status_code=400, detail="package_id is required for package promos")
        baseline = _pricing_engine.calculate_for_package(
            user_id=current_user["id"],
            package_id=req.package_id,
            quantity=req.quantity,
            promo_code=None,
        )
        subtotal_cents = int(baseline["subtotal_cents"])
    else:
        if req.subtotal_cents is None:
            raise HTTPException(status_code=400, detail="subtotal_cents is required for applies_to='all'")
        subtotal_cents = int(req.subtotal_cents)

    effect = _promo_service.validate_and_calculate(
        user_id=current_user["id"],
        code=code,
        subtotal_cents=subtotal_cents,
        applies_to=applies_to,
    )

    if not effect:
        return {"valid": False}

    promo_def = _promo_service.get_promo_by_code(code) or {}
    return {
        "valid": True,
        "code": effect.code,
        "type": promo_def.get("type"),
        "discount_percentage": promo_def.get("discount_percentage"),
        "discount_amount_cents": promo_def.get("discount_amount_cents"),
        "discount_cents": effect.discount_cents,
        "bonus_credits": effect.bonus_credits,
        "subtotal_cents": subtotal_cents,
        "total_cents": max(0, subtotal_cents - int(effect.discount_cents or 0)),
    }


@app.get("/api/promos/my-referral-code")
async def my_referral_code(current_user: dict = Depends(get_current_user)):
    from promotions.referral_service import ReferralService

    code = ReferralService().ensure_user_referral_code(current_user["id"])
    return {"referral_code": code}


@app.get("/api/promos/referral-stats")
async def referral_stats(current_user: dict = Depends(get_current_user)):
    from promotions.referral_service import ReferralService

    return ReferralService().get_referrer_stats(current_user["id"])


@app.get("/api/pricing/my-tier")
async def my_pricing_tier(current_user: dict = Depends(get_current_user)):
    tier = _pricing_engine.get_user_pricing_tier(current_user["id"])
    if not tier:
        return None
    return {
        "id": tier.id,
        "name": tier.name,
        "price_per_credit_cents": tier.price_per_credit_cents,
        "discount_percentage": tier.discount_percentage,
        "requires_approval": tier.requires_approval,
    }


@app.get("/api/pricing/tiers")
async def list_pricing_tiers(current_user: dict = Depends(get_current_user)):
    from db.payment_crud import list_pricing_tiers as _list

    tiers = []
    for r in _list(active_only=True):
        # schema: id, name, min_monthly_credits, max_monthly_credits, price_per_credit_cents,
        # discount_percentage, requires_approval, is_active, created_at, updated_at
        tiers.append({
            "id": r[0],
            "name": r[1],
            "min_monthly_credits": r[2],
            "max_monthly_credits": r[3],
            "price_per_credit_cents": r[4],
            "discount_percentage": r[5],
            "requires_approval": bool(r[6]),
            "is_active": bool(r[7]),
        })
    return tiers


@app.post("/api/admin/pricing/tiers")
async def admin_create_pricing_tier(req: PricingTierCreateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import create_pricing_tier

    tier_id = create_pricing_tier(
        name=req.name,
        min_monthly_credits=req.min_monthly_credits,
        max_monthly_credits=req.max_monthly_credits,
        price_per_credit_cents=req.price_per_credit_cents,
        discount_percentage=req.discount_percentage,
        requires_approval=req.requires_approval,
        is_active=req.is_active,
    )
    record_audit_event(
        "admin.pricing.tier_create", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="pricing_tier", target_id=str(tier_id),
        after={"name": req.name},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"id": int(tier_id)}


@app.put("/api/admin/pricing/tiers/{tier_id}")
async def admin_update_pricing_tier(tier_id: int, req: PricingTierUpdateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import update_pricing_tier

    updates = {}
    for field in ("name", "min_monthly_credits", "max_monthly_credits", "price_per_credit_cents", "discount_percentage", "requires_approval", "is_active"):
        value = getattr(req, field)
        if value is not None:
            if field in ("requires_approval", "is_active"):
                updates[field] = 1 if bool(value) else 0
            else:
                updates[field] = value

    ok = update_pricing_tier(tier_id, updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Tier not found")
    record_audit_event(
        "admin.pricing.tier_update", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="pricing_tier", target_id=str(tier_id),
        after=updates,
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


# ==================== ADMIN PROMO MANAGEMENT ====================

@app.post("/api/admin/promos")
async def admin_create_promo(req: PromoCreateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import create_promo_code

    promo_id = create_promo_code(
        code=req.code,
        promo_type=req.type,
        discount_percentage=req.discount_percentage,
        discount_amount_cents=req.discount_amount_cents,
        bonus_credits=req.bonus_credits,
        max_uses=req.max_uses,
        max_uses_per_user=req.max_uses_per_user,
        min_purchase_cents=req.min_purchase_cents,
        valid_from=req.valid_from,
        valid_until=req.valid_until,
        applies_to=req.applies_to,
        is_active=req.is_active,
    )
    record_audit_event(
        "admin.promo.create", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="promo", target_id=str(promo_id),
        after={"code": req.code, "type": req.type},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"id": int(promo_id)}


@app.get("/api/admin/promos")
async def admin_list_promos(admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import list_promo_codes

    promos = []
    for r in list_promo_codes(active_only=False):
        # schema: id, code, type, discount_percentage, discount_amount_cents, bonus_credits, max_uses, uses_count,
        # max_uses_per_user, min_purchase_cents, valid_from, valid_until, applies_to, is_active, created_at, updated_at
        promos.append({
            "id": r[0],
            "code": r[1],
            "type": r[2],
            "discount_percentage": r[3],
            "discount_amount_cents": r[4],
            "bonus_credits": r[5],
            "max_uses": r[6],
            "uses_count": r[7],
            "max_uses_per_user": r[8],
            "min_purchase_cents": r[9],
            "valid_from": r[10],
            "valid_until": r[11],
            "applies_to": r[12],
            "is_active": bool(r[13]),
            "created_at": r[14],
            "updated_at": r[15],
        })
    return {"promos": promos}


@app.put("/api/admin/promos/{promo_id}")
async def admin_update_promo(promo_id: int, req: PromoUpdateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import update_promo_code

    updates = {}
    for field in (
        "code",
        "type",
        "discount_percentage",
        "discount_amount_cents",
        "bonus_credits",
        "max_uses",
        "max_uses_per_user",
        "min_purchase_cents",
        "valid_from",
        "valid_until",
        "applies_to",
        "is_active",
    ):
        value = getattr(req, field)
        if value is not None:
            if field == "code" and isinstance(value, str):
                updates[field] = value.upper()
            elif field == "is_active":
                updates[field] = 1 if bool(value) else 0
            else:
                updates[field] = value

    ok = update_promo_code(promo_id, updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Promo not found")
    record_audit_event(
        "admin.promo.update", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="promo", target_id=str(promo_id), after=updates,
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


@app.delete("/api/admin/promos/{promo_id}")
async def admin_deactivate_promo(promo_id: int, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import deactivate_promo_code

    ok = deactivate_promo_code(promo_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Promo not found")
    record_audit_event(
        "admin.promo.delete", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="promo", target_id=str(promo_id),
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


@app.get("/api/admin/promos/{promo_id}/usage")
async def admin_promo_usage(promo_id: int, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import get_promo_usage_stats

    return get_promo_usage_stats(promo_id)


# ==================== ADMIN PACKAGES & PLANS ====================

@app.post("/api/admin/payments/packages")
async def admin_create_credit_package(req: CreditPackageCreateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import create_credit_package

    features_json = json.dumps(req.features) if req.features is not None else None
    pkg_id = create_credit_package(
        name=req.name,
        credits=req.credits,
        base_price_cents=req.base_price_cents,
        display_price_cents=req.display_price_cents,
        discount_percentage=req.discount_percentage,
        tier_id=req.tier_id,
        is_active=req.is_active,
        is_featured=req.is_featured,
        features=features_json,
    )
    record_audit_event(
        "admin.package.create", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="credit_package", target_id=str(pkg_id),
        after={"name": req.name, "credits": req.credits, "price_cents": req.display_price_cents},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"id": int(pkg_id)}


@app.get("/api/admin/payments/packages")
async def admin_list_credit_packages(admin_user: dict = Depends(get_admin_user)):
    from db.base import execute_query

    rows = execute_query("SELECT * FROM credit_packages ORDER BY display_price_cents ASC", fetch_all=True, commit=False)
    packages = []
    for row in rows or []:
        features = None
        try:
            features = json.loads(row[9]) if row[9] else None
        except Exception:
            features = None
        packages.append({
            "id": row[0],
            "name": row[1],
            "credits": row[2],
            "base_price_cents": row[3],
            "display_price_cents": row[4],
            "discount_percentage": row[5],
            "tier_id": row[6],
            "is_active": bool(row[7]),
            "is_featured": bool(row[8]),
            "features": features,
        })
    return {"packages": packages}


@app.put("/api/admin/payments/packages/{package_id}")
async def admin_update_credit_package(package_id: int, req: CreditPackageUpdateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import update_credit_package

    updates = {}
    for field in ("name", "credits", "base_price_cents", "display_price_cents", "discount_percentage", "tier_id"):
        value = getattr(req, field)
        if value is not None:
            updates[field] = value
    if req.is_active is not None:
        updates["is_active"] = 1 if bool(req.is_active) else 0
    if req.is_featured is not None:
        updates["is_featured"] = 1 if bool(req.is_featured) else 0
    if req.features is not None:
        updates["features"] = json.dumps(req.features)

    ok = update_credit_package(package_id, updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Package not found")
    record_audit_event(
        "admin.package.update", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="credit_package", target_id=str(package_id), after=updates,
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


@app.delete("/api/admin/payments/packages/{package_id}")
async def admin_delete_credit_package(package_id: int, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import delete_credit_package

    ok = delete_credit_package(package_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Package not found")
    record_audit_event(
        "admin.package.delete", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="credit_package", target_id=str(package_id),
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


@app.post("/api/admin/subscriptions/plans")
async def admin_create_subscription_plan(req: SubscriptionPlanCreateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import create_subscription_plan

    features_json = json.dumps(req.features) if req.features is not None else None
    plan_id = create_subscription_plan(
        name=req.name,
        billing_interval=req.billing_interval,
        credits_per_period=req.credits_per_period,
        base_price_cents=req.base_price_cents,
        stripe_price_id=req.stripe_price_id,
        rollover_credits=req.rollover_credits,
        max_rollover_credits=req.max_rollover_credits,
        trial_days=req.trial_days,
        is_active=req.is_active,
        is_featured=req.is_featured,
        features=features_json,
    )
    record_audit_event(
        "admin.plan.create", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="subscription_plan", target_id=str(plan_id),
        after={"name": req.name, "billing_interval": req.billing_interval},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"id": int(plan_id)}


@app.get("/api/admin/subscriptions/plans")
async def admin_list_subscription_plans(admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import list_subscription_plans

    plans = []
    for row in list_subscription_plans(active_only=False):
        features = None
        try:
            features = json.loads(row[11]) if row[11] else None
        except Exception:
            features = None
        plans.append({
            "id": row[0],
            "name": row[1],
            "billing_interval": row[2],
            "credits_per_period": row[3],
            "base_price_cents": row[4],
            "stripe_price_id": row[5],
            "rollover_credits": bool(row[6]),
            "max_rollover_credits": row[7],
            "trial_days": row[8],
            "is_active": bool(row[9]),
            "is_featured": bool(row[10]),
            "features": features,
        })
    return {"plans": plans}


@app.put("/api/admin/subscriptions/plans/{plan_id}")
async def admin_update_subscription_plan(plan_id: int, req: SubscriptionPlanUpdateRequest, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import update_subscription_plan

    updates = {}
    for field in ("name", "billing_interval", "credits_per_period", "base_price_cents", "stripe_price_id", "max_rollover_credits", "trial_days"):
        value = getattr(req, field)
        if value is not None:
            updates[field] = value
    if req.rollover_credits is not None:
        updates["rollover_credits"] = 1 if bool(req.rollover_credits) else 0
    if req.is_active is not None:
        updates["is_active"] = 1 if bool(req.is_active) else 0
    if req.is_featured is not None:
        updates["is_featured"] = 1 if bool(req.is_featured) else 0
    if req.features is not None:
        updates["features"] = json.dumps(req.features)

    ok = update_subscription_plan(plan_id, updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Plan not found")
    record_audit_event(
        "admin.plan.update", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="subscription_plan", target_id=str(plan_id), after=updates,
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}


@app.delete("/api/admin/subscriptions/plans/{plan_id}")
async def admin_delete_subscription_plan(plan_id: int, request: Request, admin_user: dict = Depends(get_admin_user)):
    from db.payment_crud import delete_subscription_plan

    ok = delete_subscription_plan(plan_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Plan not found")
    record_audit_event(
        "admin.plan.delete", request_id=request.state.request_id,
        actor_user_id=admin_user["id"], actor_username=admin_user["username"],
        target_type="subscription_plan", target_id=str(plan_id),
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"ok": True}

# ==================== ADMIN ACTIVITY / AUDIT ENDPOINTS ====================


@app.get("/api/admin/activity")
async def get_activity_feed_endpoint(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    target_type: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    admin: dict = Depends(get_admin_user),
):
    """Global activity feed with filters."""
    from audit.queries import get_activity_feed as _get_feed, get_audit_event_count
    events = _get_feed(
        actor_user_id=user_id, action=action, target_type=target_type,
        status=status, date_from=date_from, date_to=date_to,
        limit=min(limit, 200), offset=offset,
    )
    total = get_audit_event_count(
        actor_user_id=user_id, action=action, target_type=target_type,
        status=status, date_from=date_from, date_to=date_to,
    )
    return {"events": events, "total": total, "limit": limit, "offset": offset}


@app.get("/api/admin/activity/actions")
async def get_activity_actions(admin: dict = Depends(get_admin_user)):
    """Return distinct action strings for filter dropdowns."""
    from audit.queries import get_distinct_actions
    return {"actions": get_distinct_actions()}


@app.get("/api/admin/activity/export")
async def export_activity_csv_endpoint(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    admin: dict = Depends(get_admin_user),
):
    """Export activity log as CSV."""
    from audit.queries import export_activity_csv as _export
    import csv
    from io import StringIO

    rows = _export(actor_user_id=user_id, action=action, date_from=date_from, date_to=date_to)
    output = StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=activity_log.csv"},
    )


@app.get("/api/admin/users/{user_id}/timeline")
async def get_user_timeline_endpoint(
    user_id: int,
    limit: int = 100,
    offset: int = 0,
    admin: dict = Depends(get_admin_user),
):
    """User-specific activity timeline."""
    from audit.queries import get_user_timeline as _get_timeline
    events = _get_timeline(user_id=user_id, limit=limit, offset=offset)
    return {"events": events, "user_id": user_id}


@app.get("/api/admin/users/{user_id}/kpis")
async def get_user_kpis_endpoint(
    user_id: int,
    admin: dict = Depends(get_admin_user),
):
    """Per-user value KPIs."""
    from audit.queries import get_user_kpis as _kpis
    return _kpis(user_id)


@app.get("/api/admin/users/{user_id}/jobs")
async def get_user_jobs_admin(
    user_id: int,
    admin: dict = Depends(get_admin_user),
):
    """List all jobs for a specific user (admin view)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT j.job_id, j.category, j.status, j.progress, j.total_cities,
               j.credit_estimate, j.credit_charged, j.created_at, j.completed_at,
               (SELECT COUNT(*) FROM results r WHERE r.job_id = j.job_id) as result_count
        FROM jobs j WHERE j.user_id = ?
        ORDER BY j.created_at DESC
    """, (user_id,))
    jobs = [
        {
            "job_id": r[0], "category": r[1], "status": r[2], "progress": r[3],
            "total_cities": r[4], "credit_estimate": r[5], "credit_charged": r[6],
            "created_at": r[7], "completed_at": r[8], "result_count": r[9],
        }
        for r in cursor.fetchall()
    ]
    conn.close()
    return {"jobs": jobs, "user_id": user_id}


@app.get("/api/admin/jobs/{job_id}/results")
async def get_job_results_admin(
    job_id: str,
    admin: dict = Depends(get_admin_user),
):
    """View job results (admin view, no ownership check)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, business_name, phone, website, address, category, city, state, google_maps_url
        FROM results WHERE job_id = ?
    """, (job_id,))
    results = [
        {
            "id": r[0], "business_name": r[1], "phone": r[2], "website": r[3],
            "address": r[4], "category": r[5], "city": r[6], "state": r[7],
            "google_maps_url": r[8],
        }
        for r in cursor.fetchall()
    ]
    conn.close()
    return {"results": results, "job_id": job_id}


@app.get("/api/admin/audit/integrity")
async def verify_audit_integrity(
    limit: int = 1000,
    admin: dict = Depends(get_owner_user),
):
    """Verify tamper-evident hash chain integrity (owner only)."""
    from audit.retention import verify_hash_chain
    return verify_hash_chain(limit=limit)


@app.post("/api/admin/audit/archive")
async def trigger_audit_archive(
    request: Request,
    admin: dict = Depends(get_owner_user),
):
    """Manually trigger archival of old audit events (owner only)."""
    from audit.retention import archive_old_events
    count = archive_old_events(hot_days=90)
    record_audit_event(
        "admin.audit.archive", request_id=request.state.request_id,
        actor_user_id=admin["id"], actor_username=admin["username"],
        metadata={"archived_count": count},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )
    return {"archived": count}


# ==================== 2-STEP APPROVAL ENDPOINTS ====================


@app.get("/api/admin/approvals/pending")
async def list_pending_approvals(admin: dict = Depends(get_admin_user)):
    """List pending 2-step approval requests."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pa.id, pa.request_id, pa.requested_by, u.username,
               pa.action, pa.target_type, pa.target_id, pa.payload,
               pa.status, pa.expires_at, pa.created_at
        FROM pending_approvals pa
        LEFT JOIN users u ON pa.requested_by = u.id
        WHERE pa.status = 'pending' AND pa.expires_at > ?
        ORDER BY pa.created_at DESC
    """, (datetime.now().isoformat(),))
    approvals = [
        {
            "id": r[0], "request_id": r[1], "requested_by": r[2],
            "requested_by_username": r[3], "action": r[4],
            "target_type": r[5], "target_id": r[6],
            "payload": json.loads(r[7]) if r[7] else None,
            "status": r[8], "expires_at": r[9], "created_at": r[10],
        }
        for r in cursor.fetchall()
    ]
    conn.close()
    return {"approvals": approvals}


@app.post("/api/admin/approvals/{approval_id}/approve")
async def approve_pending_action(
    approval_id: int,
    request: Request,
    admin: dict = Depends(get_owner_user),
):
    """Approve a pending destructive action (owner only)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, requested_by, action, target_type, target_id, payload, status, expires_at FROM pending_approvals WHERE id = ?",
        (approval_id,),
    )
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Approval request not found")

    if row[6] != "pending":
        conn.close()
        raise HTTPException(status_code=400, detail=f"Request already {row[6]}")

    if row[7] and row[7] < datetime.now().isoformat():
        conn.close()
        raise HTTPException(status_code=400, detail="Approval request has expired")

    if row[1] == admin["id"]:
        conn.close()
        raise HTTPException(status_code=400, detail="Cannot approve your own request")

    now = datetime.now().isoformat()
    cursor.execute(
        "UPDATE pending_approvals SET status = 'approved', reviewed_by = ?, reviewed_at = ? WHERE id = ?",
        (admin["id"], now, approval_id),
    )
    conn.commit()
    conn.close()

    record_audit_event(
        "admin.approval.approve", request_id=request.state.request_id,
        actor_user_id=admin["id"], actor_username=admin["username"],
        target_type="pending_approval", target_id=str(approval_id),
        metadata={"original_action": row[2], "target_type": row[3], "target_id": row[4]},
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )

    return {"ok": True, "message": "Action approved"}


@app.post("/api/admin/approvals/{approval_id}/reject")
async def reject_pending_action(
    approval_id: int,
    request: Request,
    admin: dict = Depends(get_admin_or_above),
):
    """Reject a pending destructive action (admin or above)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, status FROM pending_approvals WHERE id = ?",
        (approval_id,),
    )
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Approval request not found")

    if row[1] != "pending":
        conn.close()
        raise HTTPException(status_code=400, detail=f"Request already {row[1]}")

    now = datetime.now().isoformat()
    cursor.execute(
        "UPDATE pending_approvals SET status = 'rejected', reviewed_by = ?, reviewed_at = ? WHERE id = ?",
        (admin["id"], now, approval_id),
    )
    conn.commit()
    conn.close()

    record_audit_event(
        "admin.approval.reject", request_id=request.state.request_id,
        actor_user_id=admin["id"], actor_username=admin["username"],
        target_type="pending_approval", target_id=str(approval_id),
        ip_address=request.state.client_ip, user_agent=request.state.user_agent,
    )

    return {"ok": True, "message": "Action rejected"}


# ==================== UTILITY ENDPOINTS ====================

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


# ==================== ANALYTICS ====================

@app.get("/api/analytics/spending")
async def user_spending_analytics(current_user: dict = Depends(get_current_user)):
    from analytics.user_analytics import get_spending_series

    return {"series": get_spending_series(current_user["id"], months=6)}


@app.get("/api/analytics/usage")
async def user_usage_analytics(current_user: dict = Depends(get_current_user)):
    from analytics.user_analytics import get_usage_summary

    return get_usage_summary(current_user["id"], days=30)


@app.get("/api/admin/analytics/revenue")
async def admin_revenue_dashboard(admin: dict = Depends(get_admin_user)):
    from analytics.revenue_analytics import get_revenue_dashboard

    return get_revenue_dashboard()


@app.get("/api/admin/analytics/mrr")
async def admin_mrr(admin: dict = Depends(get_admin_user)):
    from analytics.revenue_analytics import get_mrr_cents, get_mrr_series

    return {"mrr_cents": get_mrr_cents(), "series": get_mrr_series(months=6)}


@app.get("/api/admin/analytics/churn")
async def admin_churn(admin: dict = Depends(get_admin_user)):
    from analytics.revenue_analytics import get_churn_summary

    return get_churn_summary()


@app.get("/api/admin/analytics/top-customers")
async def admin_top_customers(admin: dict = Depends(get_admin_user)):
    from analytics.revenue_analytics import get_top_customers

    return {"customers": get_top_customers(limit=25)}


@app.get("/api/admin/analytics/conversion")
async def admin_conversion(admin: dict = Depends(get_admin_user)):
    from analytics.revenue_analytics import get_conversion_funnel

    return get_conversion_funnel()

@app.get("/api/states")
async def get_states():
    """Get list of all available states/regions for all countries"""
    return {
        "USA": list(STATES_CITIES_DATA.keys()),
        "UK": list(UK_REGIONS_DATA.keys()),
        "UAE": list(UAE_CITIES_DATA.keys()),
        "KSA": list(KSA_CITIES_DATA.keys()),
        "Australia": list(AUSTRALIA_CITIES_DATA.keys())
    }

@app.get("/api/states/{state}/cities")
async def get_cities(state: str):
    """Get list of cities for a specific state/region"""
    # Check each country's data for the state/region
    if state in STATES_CITIES_DATA:
        return {"country": "USA", "state": state, "cities": STATES_CITIES_DATA[state]}
    elif state in UK_REGIONS_DATA:
        return {"country": "UK", "region": state, "cities": UK_REGIONS_DATA[state]}
    elif state in UAE_CITIES_DATA:
        return {"country": "UAE", "emirate": state, "cities": UAE_CITIES_DATA[state]}
    elif state in KSA_CITIES_DATA:
        return {"country": "KSA", "region": state, "cities": KSA_CITIES_DATA[state]}
    elif state in AUSTRALIA_CITIES_DATA:
        return {"country": "Australia", "state": state, "cities": AUSTRALIA_CITIES_DATA[state]}
    else:
        raise HTTPException(status_code=404, detail="State or region not found")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
